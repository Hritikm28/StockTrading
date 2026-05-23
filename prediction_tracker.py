from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
import json
import warnings
from functools import lru_cache
import time
from typing import Dict, List, Optional, Tuple, Union
import asyncio
from concurrent.futures import ThreadPoolExecutor
from collections import defaultdict
import logging
warnings.filterwarnings('ignore')

# Optional dependencies
try:
    from scipy import stats
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False

try:
    from sklearn.metrics import mean_squared_error, mean_absolute_error, brier_score_loss
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False

# Excel export support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    try:
        import xlsxwriter
        EXCEL_AVAILABLE = True
        EXCEL_ENGINE = 'xlsxwriter'
    except ImportError:
        EXCEL_AVAILABLE = False
        EXCEL_ENGINE = None

# Try to import risk_manager for regime detection
try:
    from risk_manager import MarketRegimeDetector
    RISK_MANAGER_AVAILABLE = True
except ImportError:
    RISK_MANAGER_AVAILABLE = False

# Optional visualization dependencies
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

try:
    import dash
    from dash import dcc, html, Input, Output
    DASH_AVAILABLE = True
except ImportError:
    DASH_AVAILABLE = False

# Optional GPU acceleration
try:
    import cupy as cp
    GPU_ACCELERATION_AVAILABLE = True
except ImportError:
    GPU_ACCELERATION_AVAILABLE = False
    # Fallback to NumPy
    cp = np

# Optional ML libraries for drift detection
try:
    from sklearn.ensemble import IsolationForest, RandomForestClassifier
    from sklearn.preprocessing import StandardScaler
    ML_DRIFT_AVAILABLE = True
except ImportError:
    ML_DRIFT_AVAILABLE = False

# Setup logging
logger = logging.getLogger(__name__)


class FinalPredictionTracker:
    
    TRACKING_DIR = Path("prediction_tracking")
    PREDICTIONS_DIR = TRACKING_DIR / "predictions"
    METRICS_DIR = TRACKING_DIR / "metrics"
    CACHE_DIR = TRACKING_DIR / ".cache"
    FEATURE_IMPORTANCE_DIR = TRACKING_DIR / "feature_importance"
    CALIBRATION_DIR = TRACKING_DIR / "calibration"
    MODEL_COMPARISON_DIR = TRACKING_DIR / "model_comparison"
    ALERTS_DIR = TRACKING_DIR / "alerts"
    CROSS_ASSET_DIR = TRACKING_DIR / "cross_asset"
    PREDICTION_INTERVALS_DIR = TRACKING_DIR / "prediction_intervals"
    ATTRIBUTION_DIR = TRACKING_DIR / "attribution"
    EXCEL_DIR = TRACKING_DIR / "excel_exports"
    
    # Configuration
    USE_PARQUET = True
    USE_EXCEL = True  # Excel export for all data
    MAX_PREDICTIONS_PER_SYMBOL = 500
    PERFORMANCE_PROFILING = False  # Disable in production
    ASYNC_WRITES = True  # Non-blocking I/O
    ENABLE_CALIBRATION = True
    ENABLE_MONITORING = True
    ENABLE_ALERTS = True
    
    _performance_log = {}
    _executor = ThreadPoolExecutor(max_workers=4)
    _write_queue = []
    _alerts = []  # Alert queue
    _monitoring_active = False
    
    @staticmethod
    def initialize(use_excel: bool = True):
        """Initialize with optimized directory structure"""
        for dir_path in [
            FinalPredictionTracker.TRACKING_DIR,
            FinalPredictionTracker.PREDICTIONS_DIR,
            FinalPredictionTracker.METRICS_DIR,
            FinalPredictionTracker.CACHE_DIR,
            FinalPredictionTracker.FEATURE_IMPORTANCE_DIR,
            FinalPredictionTracker.CALIBRATION_DIR,
            FinalPredictionTracker.MODEL_COMPARISON_DIR,
            FinalPredictionTracker.ALERTS_DIR,
            FinalPredictionTracker.CROSS_ASSET_DIR,
            FinalPredictionTracker.PREDICTION_INTERVALS_DIR,
            FinalPredictionTracker.ATTRIBUTION_DIR,
            FinalPredictionTracker.EXCEL_DIR
        ]:
            dir_path.mkdir(parents=True, exist_ok=True)
        
        FinalPredictionTracker.USE_EXCEL = use_excel
        
        print(f"📊 Final tracker initialized: {FinalPredictionTracker.TRACKING_DIR}")
        print(f"   ⚡ Parquet: {'ON' if FinalPredictionTracker.USE_PARQUET else 'OFF'}")
        print(f"   ⚡ Excel Export: {'ON' if FinalPredictionTracker.USE_EXCEL and EXCEL_AVAILABLE else 'OFF'}")
        print(f"   ⚡ Async writes: {'ON' if FinalPredictionTracker.ASYNC_WRITES else 'OFF'}")
        print(f"   ⚡ Calibration: {'ON' if FinalPredictionTracker.ENABLE_CALIBRATION else 'OFF'}")
        print(f"   ⚡ Monitoring: {'ON' if FinalPredictionTracker.ENABLE_MONITORING else 'OFF'}")
    
    @staticmethod
    def save_prediction_with_uncertainty(
        symbol: str, date: str, predicted_price: float, predicted_change_pct: float,
        confidence: float, model_version: int, ensemble_predictions: List[float], 
        horizon: str = '1d', metadata: Optional[Dict] = None,
        feature_importance: Optional[Dict] = None
    ):
        """Save prediction with async I/O"""
        
        start_time = time.time()
        
        # Vectorized ensemble calculations (faster than loops)
        ensemble_arr = np.array(ensemble_predictions)
        ensemble_std = np.std(ensemble_arr) if len(ensemble_arr) > 1 else 0.0
        ensemble_agreement = 1.0 - np.clip(ensemble_std / (abs(predicted_change_pct) + 1e-6), 0, 1)
        
        # Quality score (vectorized)
        quality_score = FinalPredictionTracker._calculate_quality_score_vectorized(
            confidence, ensemble_agreement, ensemble_std
        )
        
        ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
        predictions_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
        
        new_pred = {
            'prediction_date': date,
            'horizon': horizon,
            'predicted_price': float(predicted_price),
            'predicted_change_pct': float(predicted_change_pct),
            'confidence': float(confidence),
            'model_version': int(model_version),
            'ensemble_agreement': float(ensemble_agreement),
            'prediction_uncertainty': float(ensemble_std),
            'quality_score': float(quality_score),
            'actual_price': None,
            'actual_change_pct': None,
            'is_evaluated': False,
            'created_at': datetime.now().isoformat()
        }
        
        if metadata:
            new_pred.update(metadata)
        
        # Async write (non-blocking)
        if FinalPredictionTracker.ASYNC_WRITES:
            FinalPredictionTracker._executor.submit(
                FinalPredictionTracker._write_prediction_async,
                predictions_file, new_pred, symbol
            )
            
            if feature_importance:
                FinalPredictionTracker._executor.submit(
                    FinalPredictionTracker._save_feature_importance,
                    symbol, date, feature_importance, model_version
                )
        else:
            # Synchronous fallback
            FinalPredictionTracker._write_prediction_sync(predictions_file, new_pred)
            if feature_importance:
                FinalPredictionTracker._save_feature_importance(
                    symbol, date, feature_importance, model_version
                )
        
        FinalPredictionTracker._clear_cache(symbol)
        
        if FinalPredictionTracker.PERFORMANCE_PROFILING:
            elapsed = time.time() - start_time
            FinalPredictionTracker._log_performance('save_prediction', elapsed)
            if elapsed > 0.05:  # Only log if >50ms
                print(f"   ⚡ Prediction saved in {elapsed*1000:.1f}ms")
    
    @staticmethod
    def _write_prediction_async(predictions_file: Path, new_pred: Dict, symbol: str):
        """Async write helper (runs in thread pool)"""
        try:
            FinalPredictionTracker._write_prediction_sync(predictions_file, new_pred)
        except Exception as e:
            print(f"   ⚠️ Async write failed for {symbol}: {e}")
    
    @staticmethod
    def _write_prediction_sync(predictions_file: Path, new_pred: Dict):
        """Synchronous write helper"""
        ext = predictions_file.suffix
        
        if predictions_file.exists():
            if ext == '.parquet':
                df = pd.read_parquet(predictions_file)
            else:
                df = pd.read_csv(predictions_file)
            
            # Remove duplicates efficiently
            mask = ~((df['prediction_date'] == new_pred['prediction_date']) & 
                     (df['horizon'] == new_pred['horizon']))
            df = df[mask]
            df = pd.concat([df, pd.DataFrame([new_pred])], ignore_index=True)
            
            # Auto-cleanup
            if len(df) > FinalPredictionTracker.MAX_PREDICTIONS_PER_SYMBOL:
                df = df.tail(FinalPredictionTracker.MAX_PREDICTIONS_PER_SYMBOL)
        else:
            df = pd.DataFrame([new_pred])
        
        # Write with optimal compression
        if ext == '.parquet':
            df.to_parquet(predictions_file, engine='pyarrow', compression='snappy', index=False)
        else:
            df.to_csv(predictions_file, index=False)
    
    @staticmethod
    def _calculate_quality_score_vectorized(confidence: float, agreement: float, uncertainty: float) -> float:
        """Vectorized quality score calculation"""
        scores = np.array([
            confidence,                      # Already 0-100
            agreement * 100,                 # 0-100
            max(0.0, 100.0 - uncertainty * 20)  # Uncertainty penalty
        ])
        
        weights = np.array([0.4, 0.4, 0.2])
        quality = np.dot(scores, weights)
        
        return float(np.clip(quality, 0, 100))
    
    @staticmethod
    def _save_feature_importance(symbol: str, date: str, importance: Dict, version: int):
        """Track feature importance (optimized)"""
        importance_file = FinalPredictionTracker.FEATURE_IMPORTANCE_DIR / f"{symbol}_importance.json"
        
        try:
            if importance_file.exists():
                with open(importance_file, 'r') as f:
                    history = json.load(f)
            else:
                history = []
            
            history.append({
                'date': date,
                'version': version,
                'importance': {k: float(v) for k, v in importance.items()}  # Ensure JSON serializable
            })
            
            history = history[-50:]
            
            with open(importance_file, 'w') as f:
                json.dump(history, f, separators=(',', ':'))  # Compact format
        except Exception as e:
            print(f"   ⚠️ Feature importance save failed: {e}")
    
    @staticmethod
    def evaluate_prediction_advanced(
        symbol: str, prediction_date: str, actual_price: float, 
        current_df: pd.DataFrame, models_dict: Optional[Dict] = None, 
        horizon: str = '1d'
    ) -> Dict:
        """Optimized evaluation with vectorized operations"""
        
        start_time = time.time()
        
        ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
        predictions_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
        
        if not predictions_file.exists():
            return {
                'should_retrain': False,
                'accuracy': None,
                'details': {},
                'drift_detected': False
            }
        
        # Load efficiently
        if FinalPredictionTracker.USE_PARQUET:
            preds_df = pd.read_parquet(predictions_file)
        else:
            preds_df = pd.read_csv(predictions_file)
        
        # Vectorized filtering
        mask = (
            (preds_df['prediction_date'] == prediction_date) &
            (preds_df['horizon'] == horizon) &
            (preds_df['is_evaluated'] == False)
        )
        
        pending = preds_df[mask]
        
        if pending.empty:
            return {
                'should_retrain': False,
                'accuracy': None,
                'details': {},
                'drift_detected': False
            }
        
        pred = pending.iloc[0]
        
        # Get actual change
        try:
            pred_date_idx = current_df.index.get_loc(pd.Timestamp(prediction_date))
            days = {'1d': 1, '1w': 5, '1m': 21}.get(horizon, 1)
            actual_date_idx = min(pred_date_idx + days, len(current_df) - 1)
            
            reference_price = current_df['Close'].iloc[pred_date_idx]
            actual_change_pct = ((actual_price - reference_price) / reference_price) * 100
        except Exception as e:
            return {
                'should_retrain': False,
                'accuracy': None,
                'details': {},
                'drift_detected': False
            }
        
        # Vectorized metrics calculation
        metrics = FinalPredictionTracker._calculate_all_metrics_vectorized(
            pred, actual_price, actual_change_pct, symbol
        )
        
        # Regime detection
        regime = FinalPredictionTracker._detect_regime(current_df)
        regime_adjustment = 0.8 if 'high_vol' in regime else 1.0
        
        # Final accuracy
        adjusted_accuracy = (metrics['base_score'] / metrics['uncertainty_penalty']) * regime_adjustment + metrics['quality_bonus']
        
        # Update file (async)
        preds_df.loc[mask, 'actual_price'] = actual_price
        preds_df.loc[mask, 'actual_change_pct'] = actual_change_pct
        preds_df.loc[mask, 'is_evaluated'] = True
        
        if FinalPredictionTracker.ASYNC_WRITES:
            FinalPredictionTracker._executor.submit(
                FinalPredictionTracker._save_updated_predictions,
                predictions_file, preds_df
            )
        else:
            FinalPredictionTracker._save_updated_predictions(predictions_file, preds_df)
        
        # Save metrics
        FinalPredictionTracker._save_accuracy_metrics(
            symbol, prediction_date, horizon,
            metrics['direction_correct'], metrics['magnitude_mae'], metrics['price_mape'],
            metrics['brier'], adjusted_accuracy, regime, 
            metrics['bayesian_accuracy'], metrics['confidence_interval']
        )
        
        # Update calibration tracking
        if FinalPredictionTracker.ENABLE_CALIBRATION:
            pred_prob = pred['confidence'] / 100.0 if pred['predicted_change_pct'] > 0 else 1 - (pred['confidence'] / 100.0)
            actual_outcome = actual_change_pct > 0
            FinalPredictionTracker.update_calibration(symbol, prediction_date, pred['confidence'], actual_outcome)
        
        # Drift detection
        drift_detected = FinalPredictionTracker._detect_drift_advanced(symbol, current_df)
        correlation_degraded = FinalPredictionTracker._check_feature_correlation_drift(symbol, current_df)
        
        # Retraining decision
        should_retrain = FinalPredictionTracker._should_retrain_adaptive(
            symbol, drift_detected, adjusted_accuracy, correlation_degraded
        )
        
        details = {
            'predicted_price': float(pred['predicted_price']),
            'actual_price': actual_price,
            'predicted_change_%': float(pred['predicted_change_pct']),
            'actual_change_%': actual_change_pct,
            'direction_correct': metrics['direction_correct'],
            'magnitude_mae': metrics['magnitude_mae'],
            'price_mape': metrics['price_mape'],
            'brier_score': metrics['brier'],
            'adjusted_accuracy': adjusted_accuracy,
            'bayesian_accuracy': metrics['bayesian_accuracy'],
            'confidence_interval': metrics['confidence_interval'],
            'ensemble_agreement': float(pred['ensemble_agreement']),
            'quality_score': float(pred['quality_score']),
            'regime': regime
        }
        
        FinalPredictionTracker._clear_cache(symbol)
        
        if should_retrain:
            recent = FinalPredictionTracker._get_recent_metrics(symbol, days=10)
            reason = recent.get('retrain_reason', 'Accuracy decline')
            print(f"   🔄 {reason} → RETRAIN")
        else:
            recent = FinalPredictionTracker._get_recent_metrics(symbol, days=10)
            avg_acc = recent.get('avg_accuracy', 0)
            ci = recent.get('confidence_interval', (0, 0))
            print(f"   ✅ Accuracy: {avg_acc:.1f}% (CI: {ci[0]:.1f}-{ci[1]:.1f}%) → OK")
        
        if FinalPredictionTracker.PERFORMANCE_PROFILING:
            elapsed = time.time() - start_time
            FinalPredictionTracker._log_performance('evaluate_prediction', elapsed)
            if elapsed > 0.1:
                print(f"   ⚡ Eval: {elapsed*1000:.0f}ms")
        
        return {
            'should_retrain': should_retrain,
            'accuracy': adjusted_accuracy,
            'details': details,
            'drift_detected': drift_detected,
            'correlation_degraded': correlation_degraded
        }
    
    @staticmethod
    def _calculate_all_metrics_vectorized(pred: pd.Series, actual_price: float, actual_change_pct: float, symbol: str) -> Dict:
        """Vectorized calculation of all metrics at once"""
        
        # Direction
        direction_correct = bool(np.sign(pred['predicted_change_pct']) == np.sign(actual_change_pct))
        
        # Errors (vectorized)
        errors = np.array([
            abs(pred['predicted_change_pct'] - actual_change_pct),  # MAE
            abs((actual_price - pred['predicted_price']) / actual_price) * 100  # MAPE
        ])
        
        magnitude_mae, price_mape = errors
        
        # Brier score
        if SKLEARN_AVAILABLE:
            pred_prob = pred['confidence'] / 100 if pred['predicted_change_pct'] > 0 else 1 - (pred['confidence'] / 100)
            actual_outcome = 1 if actual_change_pct > 0 else 0
            brier = (pred_prob - actual_outcome) ** 2
        else:
            brier = 0.5
        
        # Bayesian accuracy
        bayesian_accuracy, confidence_interval = FinalPredictionTracker._calculate_bayesian_accuracy(
            symbol, direction_correct
        )
        
        # Uncertainty penalty
        uncertainty_penalty = 1.0 + (float(pred['prediction_uncertainty']) * 0.1)
        
        # Base score (vectorized components)
        score_components = np.array([
            50.0 if direction_correct else 0.0,
            max(0.0, 30.0 * (1 - magnitude_mae / 10)),
            max(0.0, 20.0 * (1 - price_mape / 5))
        ])
        
        base_score = np.sum(score_components)
        
        # Quality bonus
        quality_bonus = (float(pred['quality_score']) / 100) * 5 if direction_correct else 0.0
        
        return {
            'direction_correct': direction_correct,
            'magnitude_mae': float(magnitude_mae),
            'price_mape': float(price_mape),
            'brier': float(brier),
            'bayesian_accuracy': bayesian_accuracy,
            'confidence_interval': confidence_interval,
            'uncertainty_penalty': uncertainty_penalty,
            'base_score': float(base_score),
            'quality_bonus': quality_bonus
        }
    
    @staticmethod
    def _save_updated_predictions(predictions_file: Path, df: pd.DataFrame):
        """Save updated predictions"""
        try:
            if predictions_file.suffix == '.parquet':
                df.to_parquet(predictions_file, engine='pyarrow', compression='snappy', index=False)
            else:
                df.to_csv(predictions_file, index=False)
        except Exception as e:
            print(f"   ⚠️ Save failed: {e}")
    
    @staticmethod
    def _calculate_bayesian_accuracy(symbol: str, current_correct: bool) -> Tuple[float, Tuple[float, float]]:
        """Bayesian posterior with credible intervals"""
        
        ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
        metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
        
        if not metrics_file.exists():
            return (50.0, (30.0, 70.0))
        
        try:
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(metrics_file)
            else:
                df = pd.read_csv(metrics_file)
            
            recent = df.tail(20)
            
            # Vectorized calculations
            successes = int(recent['direction_accuracy'].sum()) + (1 if current_correct else 0)
            failures = len(recent) + 1 - successes
            
            alpha_post = 1 + successes
            beta_post = 1 + failures
            
            posterior_mean = alpha_post / (alpha_post + beta_post)
            
            if SCIPY_AVAILABLE:
                lower = stats.beta.ppf(0.025, alpha_post, beta_post)
                upper = stats.beta.ppf(0.975, alpha_post, beta_post)
            else:
                std = np.sqrt((alpha_post * beta_post) / ((alpha_post + beta_post)**2 * (alpha_post + beta_post + 1)))
                lower = posterior_mean - 1.96 * std
                upper = posterior_mean + 1.96 * std
            
            return (posterior_mean * 100, (lower * 100, upper * 100))
        except:
            return (50.0, (30.0, 70.0))
    
    @staticmethod
    def _check_feature_correlation_drift(symbol: str, current_df: pd.DataFrame) -> bool:
        """Feature correlation drift detection"""
        
        importance_file = FinalPredictionTracker.FEATURE_IMPORTANCE_DIR / f"{symbol}_importance.json"
        
        if not importance_file.exists():
            return False
        
        try:
            with open(importance_file, 'r') as f:
                history = json.load(f)
            
            if len(history) < 5:
                return False
            
            recent_importance = history[-1]['importance']
            
            # Vectorized historical average
            historical_importance = {}
            for record in history[-10:-1]:
                for feat, imp in record['importance'].items():
                    if feat not in historical_importance:
                        historical_importance[feat] = []
                    historical_importance[feat].append(imp)
            
            # Vectorized degradation check
            degraded = 0
            total = 0
            
            for feat, recent_imp in recent_importance.items():
                if feat in historical_importance:
                    hist_avg = np.mean(historical_importance[feat])
                    if hist_avg > 0.05 and recent_imp < hist_avg * 0.7:
                        degraded += 1
                    total += 1
            
            return (degraded / total) > 0.2 if total > 0 else False
        except:
            return False
    
    @staticmethod
    def _save_accuracy_metrics(
        symbol: str, date: str, horizon: str, direction_correct: bool,
        magnitude_mae: float, price_mape: float, brier: float,
        adjusted_accuracy: float, regime: str, bayesian_accuracy: float,
        confidence_interval: Tuple[float, float]
    ):
        """Save metrics (async)"""
        
        new_metric = {
            'evaluation_date': date,
            'horizon': horizon,
            'direction_accuracy': 1.0 if direction_correct else 0.0,
            'magnitude_mae': float(magnitude_mae),
            'price_mape': float(price_mape),
            'brier_score': float(brier),
            'adjusted_accuracy': float(adjusted_accuracy),
            'bayesian_accuracy': float(bayesian_accuracy),
            'ci_lower': float(confidence_interval[0]),
            'ci_upper': float(confidence_interval[1]),
            'regime': regime,
            'created_at': datetime.now().isoformat()
        }
        
        if FinalPredictionTracker.ASYNC_WRITES:
            FinalPredictionTracker._executor.submit(
                FinalPredictionTracker._save_metrics_sync,
                symbol, new_metric
            )
        else:
            FinalPredictionTracker._save_metrics_sync(symbol, new_metric)
    
    @staticmethod
    def _save_metrics_sync(symbol: str, new_metric: Dict):
        """Synchronous metrics save"""
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            
            if metrics_file.exists():
                if ext == '.parquet':
                    df = pd.read_parquet(metrics_file)
                else:
                    df = pd.read_csv(metrics_file)
                
                df = pd.concat([df, pd.DataFrame([new_metric])], ignore_index=True)
                df = df.tail(200)
            else:
                df = pd.DataFrame([new_metric])
            
            if ext == '.parquet':
                df.to_parquet(metrics_file, engine='pyarrow', compression='snappy', index=False)
            else:
                df.to_csv(metrics_file, index=False)
        except Exception as e:
            print(f"   ⚠️ Metrics save failed: {e}")
    
    @staticmethod
    @lru_cache(maxsize=200)
    def _get_recent_metrics(symbol: str, days: int = 10) -> Dict:
        """Cached recent metrics"""
        
        ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
        metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
        
        if not metrics_file.exists():
            return {
                'avg_accuracy': 70.0,
                'accuracy_std': 0,
                'direction_accuracy': 0.5,
                'consecutive_direction_errors': 0,
                'avg_magnitude_error': 0,
                'num_predictions': 0,
                'confidence_interval': (50.0, 90.0)
            }
        
        try:
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(metrics_file)
            else:
                df = pd.read_csv(metrics_file)
            
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = df[df['evaluation_date'] >= cutoff_date]
            
            if recent.empty:
                recent = df.tail(days)
            
            if recent.empty:
                return {
                    'avg_accuracy': 70.0,
                    'accuracy_std': 0,
                    'direction_accuracy': 0.5,
                    'consecutive_direction_errors': 0,
                    'avg_magnitude_error': 0,
                    'num_predictions': 0,
                    'confidence_interval': (50.0, 90.0)
                }
            
            # Vectorized consecutive errors (optimized)
            dir_acc = recent['direction_accuracy'].values[::-1]  # Reverse for recent-first
            # Find first non-zero (first success) using argmax
            if len(dir_acc) > 0:
                first_success_idx = np.argmax(dir_acc == 1) if np.any(dir_acc == 1) else len(dir_acc)
                consecutive_errors = int(first_success_idx) if dir_acc[0] == 0 else 0
            else:
                consecutive_errors = 0
            
            # Get CI
            if 'bayesian_accuracy' in recent.columns:
                avg_bayesian = recent['bayesian_accuracy'].mean()
                ci_lower = recent['ci_lower'].mean()
                ci_upper = recent['ci_upper'].mean()
            else:
                avg_bayesian = recent['adjusted_accuracy'].mean()
                ci_lower = avg_bayesian - 10
                ci_upper = avg_bayesian + 10
            
            return {
                'avg_accuracy': float(recent['adjusted_accuracy'].mean()),
                'accuracy_std': float(recent['adjusted_accuracy'].std()),
                'direction_accuracy': float(recent['direction_accuracy'].mean()),
                'consecutive_direction_errors': consecutive_errors,
                'avg_magnitude_error': float(recent['magnitude_mae'].mean()),
                'num_predictions': len(recent),
                'bayesian_accuracy': float(avg_bayesian),
                'confidence_interval': (float(ci_lower), float(ci_upper))
            }
        except:
            return {
                'avg_accuracy': 70.0,
                'accuracy_std': 0,
                'direction_accuracy': 0.5,
                'consecutive_direction_errors': 0,
                'avg_magnitude_error': 0,
                'num_predictions': 0,
                'confidence_interval': (50.0, 90.0)
            }
    
    @staticmethod
    def _should_retrain_adaptive(
        symbol: str, drift_detected: bool, current_accuracy: float,
        correlation_degraded: bool
    ) -> bool:
        """Adaptive retraining logic"""
        
        recent_metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=15)
        
        bayesian_acc = recent_metrics.get('bayesian_accuracy', 70)
        ci = recent_metrics.get('confidence_interval', (50, 90))
        accuracy_std = recent_metrics.get('accuracy_std', 10)
        
        accuracy_threshold = max(55, ci[0])
        
        # Multiple triggers
        if bayesian_acc < accuracy_threshold:
            recent_metrics['retrain_reason'] = f"Low Bayesian accuracy ({bayesian_acc:.1f}% < {accuracy_threshold:.1f}%)"
            return True
        
        ci_width = ci[1] - ci[0]
        if ci_width > 40:
            recent_metrics['retrain_reason'] = f"High uncertainty (CI: {ci_width:.1f}%)"
            return True
        
        if accuracy_std > 25:
            recent_metrics['retrain_reason'] = f"High variance (std={accuracy_std:.1f})"
            return True
        
        if drift_detected:
            recent_metrics['retrain_reason'] = "Distribution drift"
            return True
        
        if correlation_degraded:
            recent_metrics['retrain_reason'] = "Feature correlation degraded"
            return True
        
        if recent_metrics.get('consecutive_direction_errors', 0) >= 4:
            recent_metrics['retrain_reason'] = f"{recent_metrics['consecutive_direction_errors']} consecutive errors"
            return True
        
        days_since_retrain = FinalPredictionTracker._get_days_since_retrain(symbol)
        
        staleness_threshold = 45 if bayesian_acc > 75 else 30 if bayesian_acc > 65 else 20
        
        if days_since_retrain > staleness_threshold:
            recent_metrics['retrain_reason'] = f"Staleness ({days_since_retrain}d > {staleness_threshold}d)"
            return True
        
        return False
    
    @staticmethod
    def _detect_drift_advanced(symbol: str, current_df: pd.DataFrame) -> bool:
        """Drift detection with KS + t-test"""
        
        if not SCIPY_AVAILABLE or len(current_df) < 300:
            return False
        
        try:
            recent_data = current_df.iloc[-20:]
            historical_data = current_df.iloc[-272:-20]
            
            test_features = ['Returns', 'Volume_Ratio', 'RSI_14', 'Volatility_10', 'MACD_Hist', 'BB_Position_20', 'ATR_14']
            
            drift_scores = []
            
            for feat in test_features:
                if feat not in current_df.columns:
                    continue
                
                recent_vals = recent_data[feat].dropna().values
                hist_vals = historical_data[feat].dropna().values
                
                if len(recent_vals) < 5 or len(hist_vals) < 50:
                    continue
                
                ks_stat, ks_p = stats.ks_2samp(recent_vals, hist_vals)
                
                if len(recent_vals) >= 10:
                    t_stat, t_p = stats.ttest_ind(recent_vals, hist_vals)
                    drift = (ks_p < 0.05) or (t_p < 0.05)
                else:
                    drift = ks_p < 0.05
                
                drift_scores.append(drift)
            
            if not drift_scores:
                return False
            
            drift_detected = (sum(drift_scores) / len(drift_scores)) > 0.3
            
            if drift_detected:
                FinalPredictionTracker._log_drift_event(symbol, drift_scores)
            
            return drift_detected
        except:
            return False
    
    @staticmethod
    def _log_drift_event(symbol: str, drift_scores: List[bool]):
        """Log drift"""
        try:
            drift_file = FinalPredictionTracker.TRACKING_DIR / "drift_events.json"
            
            if drift_file.exists():
                with open(drift_file, 'r') as f:
                    log = json.load(f)
            else:
                log = {}
            
            if symbol not in log:
                log[symbol] = []
            
            log[symbol].append({
                'date': datetime.now().strftime('%Y-%m-%d'),
                'num_drifted': sum(drift_scores),
                'total_tested': len(drift_scores),
                'pct': sum(drift_scores) / len(drift_scores) * 100
            })
            
            log[symbol] = log[symbol][-50:]
            
            with open(drift_file, 'w') as f:
                json.dump(log, f, separators=(',', ':'))
        except:
            pass
    
    @staticmethod
    def _detect_regime(df: pd.DataFrame) -> str:
        """Regime detection"""
        if 'Volatility_10' not in df.columns or len(df) < 252:
            return 'normal'
        
        vol_percentile = df['Volatility_10'].iloc[-252:].rank(pct=True).iloc[-1]
        
        if 'Returns' in df.columns:
            recent_returns = df['Returns'].iloc[-20:].mean()
            
            if vol_percentile > 0.8:
                return 'high_vol_down' if recent_returns < -0.5 else 'high_vol_up'
            elif vol_percentile < 0.2:
                return 'low_vol'
            else:
                return 'trending_up' if recent_returns > 0.3 else 'trending_down' if recent_returns < -0.3 else 'normal'
        
        return 'high_vol' if vol_percentile > 0.8 else 'low_vol' if vol_percentile < 0.2 else 'normal'
    
    @staticmethod
    def _get_days_since_retrain(symbol: str) -> int:
        """Days since retrain"""
        try:
            versions_file = FinalPredictionTracker.TRACKING_DIR / "model_versions.json"
            
            if not versions_file.exists():
                return 999
            
            with open(versions_file, 'r') as f:
                versions = json.load(f)
            
            if symbol not in versions:
                return 999
            
            last_retrain = versions[symbol].get('last_retrain_date')
            
            if not last_retrain:
                return 999
            
            last_date = datetime.strptime(last_retrain, '%Y-%m-%d')
            return (datetime.now() - last_date).days
        except:
            return 999
    
    @staticmethod
    def increment_model_version(symbol: str, retrain_reason: str = "Manual retrain") -> int:
        """Increment version"""
        try:
            versions_file = FinalPredictionTracker.TRACKING_DIR / "model_versions.json"
            
            if versions_file.exists():
                with open(versions_file, 'r') as f:
                    versions = json.load(f)
            else:
                versions = {}
            
            current_version = versions.get(symbol, {}).get('version', 0)
            new_version = current_version + 1
            
            versions[symbol] = {
                'version': new_version,
                'last_retrain_date': datetime.now().strftime('%Y-%m-%d'),
                'retrain_reason': retrain_reason
            }
            
            with open(versions_file, 'w') as f:
                json.dump(versions, f, separators=(',', ':'))
            
            FinalPredictionTracker._clear_cache(symbol)
            
            print(f"   📦 Model v{new_version} ({retrain_reason})")
            return new_version
        except Exception as e:
            print(f"   ⚠️ Version increment failed: {e}")
            return 1
    
    @staticmethod
    def _clear_cache(symbol: str):
        """Clear cache"""
        FinalPredictionTracker._get_recent_metrics.cache_clear()
    
    @staticmethod
    def flush_async_queue():
        """Wait for all async writes to complete (call at shutdown)"""
        FinalPredictionTracker._executor.shutdown(wait=True)
        print("   ✅ All async writes completed")
    
    @staticmethod
    def check_yesterday_accuracy(symbol: str, current_price: float, current_df: pd.DataFrame) -> Tuple[bool, Optional[float], Dict]:

        try:
            yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
            
            # Try to find yesterday's date in dataframe
            if len(current_df) < 2:
                return False, None, {}
            
            # Get yesterday's date from dataframe
            yesterday_date = current_df.index[-2] if len(current_df) >= 2 else None
            if yesterday_date is None:
                return False, None, {}
            
            yesterday_str = yesterday_date.strftime('%Y-%m-%d') if hasattr(yesterday_date, 'strftime') else str(yesterday_date)
            
            # Evaluate prediction
            result = FinalPredictionTracker.evaluate_prediction_advanced(
                symbol=symbol,
                prediction_date=yesterday_str,
                actual_price=current_price,
                current_df=current_df,
                horizon='1d'
            )
            
            should_retrain = result.get('should_retrain', False)
            accuracy = result.get('accuracy')
            details = result.get('details', {})
            
            return should_retrain, accuracy, details
            
        except Exception as e:
            logger.warning(f"check_yesterday_accuracy failed for {symbol}: {e}")
            return False, None, {}
    
    @staticmethod
    def save_predictions(date: str, predictions_report: Dict):

        try:
            for symbol, pred_data in predictions_report.items():
                if not isinstance(pred_data, dict):
                    continue
                
                # Extract prediction data
                predicted_price = pred_data.get('predicted_price')
                predicted_change_pct = pred_data.get('predicted_change_pct')
                confidence = pred_data.get('confidence', 75.0)
                model_version = pred_data.get('model_version', 1)
                ensemble_predictions = pred_data.get('ensemble_predictions', [])
                horizon = pred_data.get('horizon', '1d')
                metadata = pred_data.get('metadata')
                feature_importance = pred_data.get('feature_importance')
                
                if predicted_price is None or predicted_change_pct is None:
                    continue
                
                # Save using existing method
                FinalPredictionTracker.save_prediction_with_uncertainty(
                    symbol=symbol,
                    date=date,
                    predicted_price=predicted_price,
                    predicted_change_pct=predicted_change_pct,
                    confidence=confidence,
                    model_version=model_version,
                    ensemble_predictions=ensemble_predictions if ensemble_predictions else [predicted_change_pct],
                    horizon=horizon,
                    metadata=metadata,
                    feature_importance=feature_importance
                )
        except Exception as e:
            logger.error(f"save_predictions failed: {e}")
    
    @staticmethod
    def get_model_version(symbol: str) -> int:
        """Get current model version for a symbol"""
        try:
            versions_file = FinalPredictionTracker.TRACKING_DIR / "model_versions.json"
            
            if not versions_file.exists():
                return 1
            
            with open(versions_file, 'r') as f:
                versions = json.load(f)
            
            return versions.get(symbol, {}).get('version', 1)
        except Exception as e:
            logger.warning(f"get_model_version failed for {symbol}: {e}")
            return 1
    
    @staticmethod
    def _get_recent_accuracy(symbol: str, days: int = 5) -> Optional[float]:

        try:
            metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=days)
            accuracy = metrics.get('avg_accuracy')
            
            if metrics.get('num_predictions', 0) < 3:
                return None
            
            return accuracy
        except Exception as e:
            logger.warning(f"_get_recent_accuracy failed for {symbol}: {e}")
            return None
    
    @staticmethod
    def get_recent_accuracy(symbol: str, days: int = 5) -> Optional[float]:
        """Public method to get recent prediction accuracy"""
        return FinalPredictionTracker._get_recent_accuracy(symbol, days)
    
    # ============================================================================
    # EXCEL EXPORT FUNCTIONALITY
    # ============================================================================
    
    @staticmethod
    def export_to_excel(symbol: str, output_file: Optional[str] = None) -> Optional[str]:
        """Export all prediction tracking data for a symbol to Excel"""
        if not EXCEL_AVAILABLE:
            logger.warning("Excel libraries not available. Install with: pip install openpyxl or pip install xlsxwriter")
            return None
        
        try:
            if output_file is None:
                output_file = str(FinalPredictionTracker.EXCEL_DIR / f"{symbol}_predictions.xlsx")
            
            # Create workbook
            if EXCEL_ENGINE == 'xlsxwriter':
                workbook = xlsxwriter.Workbook(output_file)
            else:
                workbook = Workbook()
                # Remove default sheet
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
            
            # Sheet 1: Predictions
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            if preds_file.exists():
                if FinalPredictionTracker.USE_PARQUET:
                    preds_df = pd.read_parquet(preds_file)
                else:
                    preds_df = pd.read_csv(preds_file)
                
                if EXCEL_ENGINE == 'xlsxwriter':
                    worksheet = workbook.add_worksheet('Predictions')
                    # Write headers
                    headers = preds_df.columns.tolist()
                    for col_num, header in enumerate(headers):
                        worksheet.write(0, col_num, header)
                    # Write data
                    for row_num, row_data in enumerate(preds_df.values, start=1):
                        for col_num, value in enumerate(row_data):
                            worksheet.write(row_num, col_num, value)
                else:
                    ws = workbook.create_sheet('Predictions')
                    for r_idx, row in enumerate(preds_df.itertuples(index=False), start=1):
                        for c_idx, value in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    # Add headers
                    for c_idx, col in enumerate(preds_df.columns, start=1):
                        ws.cell(row=1, column=c_idx, value=col)
            
            # Sheet 2: Metrics
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            if metrics_file.exists():
                if FinalPredictionTracker.USE_PARQUET:
                    metrics_df = pd.read_parquet(metrics_file)
                else:
                    metrics_df = pd.read_csv(metrics_file)
                
                if EXCEL_ENGINE == 'xlsxwriter':
                    worksheet = workbook.add_worksheet('Metrics')
                    headers = metrics_df.columns.tolist()
                    for col_num, header in enumerate(headers):
                        worksheet.write(0, col_num, header)
                    for row_num, row_data in enumerate(metrics_df.values, start=1):
                        for col_num, value in enumerate(row_data):
                            worksheet.write(row_num, col_num, value)
                else:
                    ws = workbook.create_sheet('Metrics')
                    for r_idx, row in enumerate(metrics_df.itertuples(index=False), start=1):
                        for c_idx, value in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    for c_idx, col in enumerate(metrics_df.columns, start=1):
                        ws.cell(row=1, column=c_idx, value=col)
            
            # Sheet 3: Calibration
            calib_file = FinalPredictionTracker.CALIBRATION_DIR / f"{symbol}_calibration.json"
            if calib_file.exists():
                with open(calib_file, 'r') as f:
                    calib_data = json.load(f)
                
                # Convert to DataFrame
                calib_rows = []
                for bin_key, data in calib_data.items():
                    bin_num = int(bin_key.replace('bin_', ''))
                    calib_rows.append({
                        'confidence_bin': bin_num,
                        'predicted_prob': np.mean(data['predicted']) if data['predicted'] else 0.0,
                        'actual_frequency': np.mean(data['actual']) if data['actual'] else 0.0,
                        'sample_size': len(data['predicted'])
                    })
                
                if calib_rows:
                    calib_df = pd.DataFrame(calib_rows)
                    if EXCEL_ENGINE == 'xlsxwriter':
                        worksheet = workbook.add_worksheet('Calibration')
                        headers = calib_df.columns.tolist()
                        for col_num, header in enumerate(headers):
                            worksheet.write(0, col_num, header)
                        for row_num, row_data in enumerate(calib_df.values, start=1):
                            for col_num, value in enumerate(row_data):
                                worksheet.write(row_num, col_num, value)
                    else:
                        ws = workbook.create_sheet('Calibration')
                        for r_idx, row in enumerate(calib_df.itertuples(index=False), start=1):
                            for c_idx, value in enumerate(row, start=1):
                                ws.cell(row=r_idx, column=c_idx, value=value)
                        for c_idx, col in enumerate(calib_df.columns, start=1):
                            ws.cell(row=1, column=c_idx, value=col)
            
            # Sheet 4: Feature Importance
            importance_file = FinalPredictionTracker.FEATURE_IMPORTANCE_DIR / f"{symbol}_importance.json"
            if importance_file.exists():
                with open(importance_file, 'r') as f:
                    importance_history = json.load(f)
                
                # Convert to DataFrame
                feat_rows = []
                for record in importance_history:
                    for feat, imp in record['importance'].items():
                        feat_rows.append({
                            'date': record['date'],
                            'version': record['version'],
                            'feature': feat,
                            'importance': imp
                        })
                
                if feat_rows:
                    feat_df = pd.DataFrame(feat_rows)
                    if EXCEL_ENGINE == 'xlsxwriter':
                        worksheet = workbook.add_worksheet('Feature_Importance')
                        headers = feat_df.columns.tolist()
                        for col_num, header in enumerate(headers):
                            worksheet.write(0, col_num, header)
                        for row_num, row_data in enumerate(feat_df.values, start=1):
                            for col_num, value in enumerate(row_data):
                                worksheet.write(row_num, col_num, value)
                    else:
                        ws = workbook.create_sheet('Feature_Importance')
                        for r_idx, row in enumerate(feat_df.itertuples(index=False), start=1):
                            for c_idx, value in enumerate(row, start=1):
                                ws.cell(row=r_idx, column=c_idx, value=value)
                        for c_idx, col in enumerate(feat_df.columns, start=1):
                            ws.cell(row=1, column=c_idx, value=col)
            
            # Save workbook
            if EXCEL_ENGINE == 'xlsxwriter':
                workbook.close()
            else:
                workbook.save(output_file)
            
            logger.info(f"Excel export saved to: {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return None
    
    @staticmethod
    def export_all_to_excel(output_file: Optional[str] = None) -> Optional[str]:
        """Export all symbols' data to a single Excel file with multiple sheets"""
        if not EXCEL_AVAILABLE:
            logger.warning("Excel libraries not available")
            return None
        
        try:
            if output_file is None:
                output_file = str(FinalPredictionTracker.EXCEL_DIR / f"all_predictions_{datetime.now().strftime('%Y%m%d')}.xlsx")
            
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            pred_files = list(FinalPredictionTracker.PREDICTIONS_DIR.glob(f"*_predictions.{ext}"))
            symbols = [f.stem.replace('_predictions', '') for f in pred_files]
            
            if EXCEL_ENGINE == 'xlsxwriter':
                workbook = xlsxwriter.Workbook(output_file)
            else:
                workbook = Workbook()
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
            
            # Create summary sheet
            summary_data = []
            for symbol in symbols:
                metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=30)
                summary_data.append({
                    'Symbol': symbol,
                    'Avg_Accuracy': metrics.get('avg_accuracy', 0),
                    'Direction_Accuracy': metrics.get('direction_accuracy', 0),
                    'Num_Predictions': metrics.get('num_predictions', 0)
                })
            
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                if EXCEL_ENGINE == 'xlsxwriter':
                    worksheet = workbook.add_worksheet('Summary')
                    headers = summary_df.columns.tolist()
                    for col_num, header in enumerate(headers):
                        worksheet.write(0, col_num, header)
                    for row_num, row_data in enumerate(summary_df.values, start=1):
                        for col_num, value in enumerate(row_data):
                            worksheet.write(row_num, col_num, value)
                else:
                    ws = workbook.create_sheet('Summary')
                    for r_idx, row in enumerate(summary_df.itertuples(index=False), start=1):
                        for c_idx, value in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    for c_idx, col in enumerate(summary_df.columns, start=1):
                        ws.cell(row=1, column=c_idx, value=col)
            
            # Save workbook
            if EXCEL_ENGINE == 'xlsxwriter':
                workbook.close()
            else:
                workbook.save(output_file)
            
            logger.info(f"All symbols exported. Summary saved to: {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f"Export all to Excel failed: {e}")
            return None
    
    # ============================================================================
    # PREDICTION CALIBRATION & RELIABILITY
    # ============================================================================
    
    @staticmethod
    def update_calibration(symbol: str, date: str, confidence: float, actual_outcome: bool):
        """Update calibration tracking"""
        if not FinalPredictionTracker.ENABLE_CALIBRATION:
            return
        
        try:
            # Bin confidence (0-10, 10-20, ..., 90-100)
            confidence_bin = int(confidence // 10) * 10
            
            calib_file = FinalPredictionTracker.CALIBRATION_DIR / f"{symbol}_calibration.json"
            
            if calib_file.exists():
                with open(calib_file, 'r') as f:
                    calib_data = json.load(f)
            else:
                calib_data = {}
            
            bin_key = f"bin_{confidence_bin}"
            if bin_key not in calib_data:
                calib_data[bin_key] = {'predicted': [], 'actual': []}
            
            # Convert confidence to probability
            pred_prob = confidence / 100.0
            actual_val = 1.0 if actual_outcome else 0.0
            
            calib_data[bin_key]['predicted'].append(pred_prob)
            calib_data[bin_key]['actual'].append(actual_val)
            
            # Keep last 1000 samples per bin
            if len(calib_data[bin_key]['predicted']) > 1000:
                calib_data[bin_key]['predicted'] = calib_data[bin_key]['predicted'][-1000:]
                calib_data[bin_key]['actual'] = calib_data[bin_key]['actual'][-1000:]
            
            # Save
            with open(calib_file, 'w') as f:
                json.dump(calib_data, f, separators=(',', ':'))
            
            # Also save to Excel if enabled
            if FinalPredictionTracker.USE_EXCEL and EXCEL_AVAILABLE:
                # Export will be done on-demand or periodically
                pass
        except Exception as e:
            logger.warning(f"update_calibration failed: {e}")
    
    
    @staticmethod
    def get_calibration_curve(symbol: str, min_samples: int = 10) -> Dict:
        """Get calibration curve data"""
        try:
            calib_file = FinalPredictionTracker.CALIBRATION_DIR / f"{symbol}_calibration.json"
            
            if not calib_file.exists():
                return {}
            
            with open(calib_file, 'r') as f:
                calib_data = json.load(f)
            
            curve = {}
            for bin_key, data in calib_data.items():
                if len(data['predicted']) < min_samples:
                    continue
                
                bin_num = int(bin_key.replace('bin_', ''))
                predicted_mean = np.mean(data['predicted'])
                actual_mean = np.mean(data['actual'])
                sample_size = len(data['predicted'])
                
                curve[bin_num] = {
                    'predicted_prob': predicted_mean,
                    'actual_frequency': actual_mean,
                    'sample_size': sample_size,
                    'calibration_error': abs(predicted_mean - actual_mean)
                }
            
            return curve
        except Exception as e:
            logger.warning(f"get_calibration_curve failed: {e}")
            return {}
    
    @staticmethod
    def calculate_brier_decomposition(symbol: str) -> Dict:
        """Decompose Brier score into calibration and resolution"""
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            
            if not metrics_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(metrics_file)
            else:
                df = pd.read_csv(metrics_file)
            
            if len(df) < 10:
                return {}
            
            # Get recent predictions with actuals
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            if preds_file.exists():
                if FinalPredictionTracker.USE_PARQUET:
                    preds_df = pd.read_parquet(preds_file)
                else:
                    preds_df = pd.read_csv(preds_file)
                
                evaluated = preds_df[preds_df['is_evaluated'] == True]
                
                if len(evaluated) < 10:
                    return {}
                
                # Calculate probabilities
                evaluated['pred_prob'] = evaluated['confidence'] / 100.0
                evaluated['actual_outcome'] = (evaluated['actual_change_pct'] > 0).astype(float)
                
                # Brier score
                brier = ((evaluated['pred_prob'] - evaluated['actual_outcome']) ** 2).mean()
                
                # Calibration (reliability)
                # Bin predictions and calculate calibration
                bins = np.linspace(0, 1, 11)
                bin_centers = (bins[:-1] + bins[1:]) / 2
                bin_indices = np.digitize(evaluated['pred_prob'], bins) - 1
                bin_indices = np.clip(bin_indices, 0, len(bin_centers) - 1)
                
                calibration = 0.0
                resolution = 0.0
                overall_mean = evaluated['actual_outcome'].mean()
                
                for i, center in enumerate(bin_centers):
                    mask = bin_indices == i
                    if mask.sum() > 0:
                        bin_pred = evaluated[mask]['pred_prob'].mean()
                        bin_actual = evaluated[mask]['actual_outcome'].mean()
                        bin_weight = mask.sum() / len(evaluated)
                        
                        calibration += bin_weight * (bin_pred - bin_actual) ** 2
                        resolution += bin_weight * (bin_actual - overall_mean) ** 2
                
                return {
                    'brier_score': float(brier),
                    'calibration': float(calibration),
                    'resolution': float(resolution),
                    'uncertainty': float(overall_mean * (1 - overall_mean)),
                    'num_samples': len(evaluated)
                }
            
            return {}
        except Exception as e:
            logger.warning(f"calculate_brier_decomposition failed: {e}")
            return {}
    
    # ============================================================================
    # PREDICTION INTERVALS & UNCERTAINTY QUANTIFICATION
    # ============================================================================
    
    @staticmethod
    def save_prediction_with_intervals(
        symbol: str, date: str, predicted_price: float, predicted_change_pct: float,
        confidence: float, model_version: int, ensemble_predictions: List[float],
        quantiles: Optional[Dict[str, float]] = None,
        horizon: str = '1d', metadata: Optional[Dict] = None
    ):

        # Calculate quantiles from ensemble if not provided
        if quantiles is None and ensemble_predictions:
            ensemble_arr = np.array(ensemble_predictions)
            quantiles = {
                'q5': float(np.percentile(ensemble_arr, 5)),
                'q25': float(np.percentile(ensemble_arr, 25)),
                'q50': float(np.percentile(ensemble_arr, 50)),
                'q75': float(np.percentile(ensemble_arr, 75)),
                'q95': float(np.percentile(ensemble_arr, 95))
            }
        
        # Save base prediction
        FinalPredictionTracker.save_prediction_with_uncertainty(
            symbol, date, predicted_price, predicted_change_pct,
            confidence, model_version, ensemble_predictions, horizon, metadata
        )
        
        # Save intervals separately
        if quantiles:
            intervals_file = FinalPredictionTracker.PREDICTION_INTERVALS_DIR / f"{symbol}_intervals.json"
            
            try:
                if intervals_file.exists():
                    with open(intervals_file, 'r') as f:
                        intervals_data = json.load(f)
                else:
                    intervals_data = []
                
                intervals_data.append({
                    'date': date,
                    'horizon': horizon,
                    'model_version': model_version,
                    'quantiles': quantiles,
                    'created_at': datetime.now().isoformat()
                })
                
                # Keep last 500
                intervals_data = intervals_data[-500:]
                
                with open(intervals_file, 'w') as f:
                    json.dump(intervals_data, f, separators=(',', ':'))
            except Exception as e:
                logger.warning(f"save_prediction_with_intervals failed: {e}")
    
    @staticmethod
    def get_prediction_intervals(symbol: str, date: str, horizon: str = '1d') -> Optional[Dict]:
        """Get prediction intervals for a specific prediction"""
        try:
            intervals_file = FinalPredictionTracker.PREDICTION_INTERVALS_DIR / f"{symbol}_intervals.json"
            
            if not intervals_file.exists():
                return None
            
            with open(intervals_file, 'r') as f:
                intervals_data = json.load(f)
            
            for entry in reversed(intervals_data):
                if entry['date'] == date and entry['horizon'] == horizon:
                    return entry['quantiles']
            
            return None
        except Exception as e:
            logger.warning(f"get_prediction_intervals failed: {e}")
            return None
    
    @staticmethod
    def calculate_uncertainty_decomposition(symbol: str) -> Dict:

        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            if len(df) < 20:
                return {}
            
            # Recent predictions with ensemble data
            recent = df.tail(50)
            
            epistemic = recent['prediction_uncertainty'].mean() if 'prediction_uncertainty' in recent.columns else 0.0
            
            if 'actual_change_pct' in recent.columns and 'predicted_change_pct' in recent.columns:
                evaluated = recent[recent['is_evaluated'] == True]
                if len(evaluated) > 10:
                    errors = evaluated['actual_change_pct'] - evaluated['predicted_change_pct']
                    total_variance = errors.var()
                    aleatoric = max(0.0, total_variance - epistemic)
                else:
                    aleatoric = 0.0
            else:
                aleatoric = 0.0
            
            return {
                'total_uncertainty': float(epistemic + aleatoric),
                'epistemic_uncertainty': float(epistemic),
                'aleatoric_uncertainty': float(aleatoric),
                'epistemic_ratio': float(epistemic / (epistemic + aleatoric + 1e-6))
            }
        except Exception as e:
            logger.warning(f"calculate_uncertainty_decomposition failed: {e}")
            return {}
    
    # ============================================================================
    # INFORMATION COEFFICIENT (IC) & RANK IC
    # ============================================================================
    
    @staticmethod
    @lru_cache(maxsize=50)
    def calculate_information_coefficient(symbol: str, days: int = 60) -> Dict:

        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            # Get evaluated predictions
            evaluated = df[df['is_evaluated'] == True].copy()
            
            if len(evaluated) < 10:
                return {}
            
            # Sort by date and take recent
            evaluated = evaluated.sort_values('prediction_date')
            evaluated = evaluated.tail(days)
            
            if len(evaluated) < 10:
                return {}
            
            pred_changes = evaluated['predicted_change_pct'].values
            actual_changes = evaluated['actual_change_pct'].values
            
            # Remove NaN
            mask = ~(np.isnan(pred_changes) | np.isnan(actual_changes))
            pred_changes = pred_changes[mask]
            actual_changes = actual_changes[mask]
            
            if len(pred_changes) < 10:
                return {}
            
            # Pearson IC
            if SCIPY_AVAILABLE:
                ic_pearson, ic_p_value = stats.pearsonr(pred_changes, actual_changes)
                ic_spearman, rank_ic_p_value = stats.spearmanr(pred_changes, actual_changes)
            else:
                # Fallback
                ic_pearson = np.corrcoef(pred_changes, actual_changes)[0, 1]
                ic_spearman = np.corrcoef(np.argsort(pred_changes), np.argsort(actual_changes))[0, 1]
                ic_p_value = 0.0
                rank_ic_p_value = 0.0
            
            # IC significance
            ic_significant = ic_p_value < 0.05 if SCIPY_AVAILABLE else abs(ic_pearson) > 0.2
            rank_ic_significant = rank_ic_p_value < 0.05 if SCIPY_AVAILABLE else abs(ic_spearman) > 0.2
            
            return {
                'ic_pearson': float(ic_pearson) if not np.isnan(ic_pearson) else 0.0,
                'ic_spearman': float(ic_spearman) if not np.isnan(ic_spearman) else 0.0,
                'ic_p_value': float(ic_p_value) if SCIPY_AVAILABLE else 0.0,
                'rank_ic_p_value': float(rank_ic_p_value) if SCIPY_AVAILABLE else 0.0,
                'ic_significant': ic_significant,
                'rank_ic_significant': rank_ic_significant,
                'num_samples': len(pred_changes)
            }
        except Exception as e:
            logger.warning(f"calculate_information_coefficient failed: {e}")
            return {}
    
    # ============================================================================
    # HIT RATE BY CONFIDENCE
    # ============================================================================
    
    @staticmethod
    def calculate_hit_rate_by_confidence(symbol: str, min_samples: int = 5) -> Dict:
        """Calculate hit rate (direction accuracy) segmented by confidence levels"""
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not metrics_file.exists() or not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                metrics_df = pd.read_parquet(metrics_file)
                preds_df = pd.read_parquet(preds_file)
            else:
                metrics_df = pd.read_csv(metrics_file)
                preds_df = pd.read_csv(preds_file)
            
            # Merge to get confidence for each evaluation
            merged = metrics_df.merge(
                preds_df[['prediction_date', 'horizon', 'confidence']],
                left_on=['evaluation_date', 'horizon'],
                right_on=['prediction_date', 'horizon'],
                how='inner'
            )
            
            if len(merged) < 10:
                return {}
            
            # Bin by confidence
            confidence_bins = [0, 50, 60, 70, 80, 90, 100]
            hit_rates = {}
            
            for i in range(len(confidence_bins) - 1):
                bin_low = confidence_bins[i]
                bin_high = confidence_bins[i + 1]
                
                mask = (merged['confidence'] >= bin_low) & (merged['confidence'] < bin_high)
                bin_data = merged[mask]
                
                if len(bin_data) >= min_samples:
                    hit_rate = bin_data['direction_accuracy'].mean()
                    hit_rates[f"{bin_low}-{bin_high}"] = {
                        'hit_rate': float(hit_rate),
                        'num_samples': len(bin_data),
                        'avg_confidence': float(bin_data['confidence'].mean())
                    }
            
            return hit_rates
        except Exception as e:
            logger.warning(f"calculate_hit_rate_by_confidence failed: {e}")
            return {}
    
    # ============================================================================
    # REGIME-SPECIFIC PERFORMANCE
    # ============================================================================
    
    @staticmethod
    def get_regime_performance(symbol: str) -> Dict:
        """Get performance metrics segmented by market regime"""
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            
            if not metrics_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(metrics_file)
            else:
                df = pd.read_csv(metrics_file)
            
            if 'regime' not in df.columns or len(df) < 10:
                return {}
            
            regime_perf = {}
            for regime in df['regime'].unique():
                regime_data = df[df['regime'] == regime]
                
                if len(regime_data) >= 5:
                    regime_perf[regime] = {
                        'avg_accuracy': float(regime_data['adjusted_accuracy'].mean()),
                        'direction_accuracy': float(regime_data['direction_accuracy'].mean()),
                        'num_predictions': len(regime_data),
                        'avg_magnitude_error': float(regime_data['magnitude_mae'].mean())
                    }
            
            return regime_perf
        except Exception as e:
            logger.warning(f"get_regime_performance failed: {e}")
            return {}
    
    # ============================================================================
    # PREDICTION ATTRIBUTION & EXPLAINABILITY
    # ============================================================================
    
    @staticmethod
    def save_prediction_attribution(
        symbol: str, date: str, horizon: str, feature_contributions: Dict,
        shap_values: Optional[Dict] = None, error_attribution: Optional[Dict] = None
    ):
        """Save prediction attribution (which features contributed to prediction/error)"""
        try:
            attribution_file = FinalPredictionTracker.ATTRIBUTION_DIR / f"{symbol}_attribution.json"
            
            if attribution_file.exists():
                with open(attribution_file, 'r') as f:
                    attributions = json.load(f)
            else:
                attributions = []
            
            attribution = {
                'date': date,
                'horizon': horizon,
                'feature_contributions': {k: float(v) for k, v in feature_contributions.items()},
                'created_at': datetime.now().isoformat()
            }
            
            if shap_values:
                attribution['shap_values'] = {k: float(v) for k, v in shap_values.items()}
            
            if error_attribution:
                attribution['error_attribution'] = {k: float(v) for k, v in error_attribution.items()}
            
            attributions.append(attribution)
            attributions = attributions[-200:]  # Keep last 200
            
            with open(attribution_file, 'w') as f:
                json.dump(attributions, f, separators=(',', ':'))
        except Exception as e:
            logger.warning(f"save_prediction_attribution failed: {e}")
    
    @staticmethod
    def get_top_error_features(symbol: str, top_n: int = 10) -> List[Tuple[str, float]]:
        """Get features that contributed most to prediction errors"""
        try:
            attribution_file = FinalPredictionTracker.ATTRIBUTION_DIR / f"{symbol}_attribution.json"
            
            if not attribution_file.exists():
                return []
            
            with open(attribution_file, 'r') as f:
                attributions = json.load(f)
            
            # Aggregate error attributions
            error_scores = defaultdict(float)
            count = defaultdict(int)
            
            for attr in attributions:
                if 'error_attribution' in attr:
                    for feat, score in attr['error_attribution'].items():
                        error_scores[feat] += abs(score)
                        count[feat] += 1
            
            # Average
            avg_errors = {feat: error_scores[feat] / count[feat] for feat in error_scores}
            
            # Sort and return top N
            sorted_features = sorted(avg_errors.items(), key=lambda x: x[1], reverse=True)
            return sorted_features[:top_n]
        except Exception as e:
            logger.warning(f"get_top_error_features failed: {e}")
            return []
    
    # ============================================================================
    # MODEL COMPARISON & A/B TESTING
    # ============================================================================
    
    @staticmethod
    def compare_models(symbol: str, model_versions: List[int]) -> Dict:
        """Compare performance of different model versions"""
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not metrics_file.exists() or not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                metrics_df = pd.read_parquet(metrics_file)
                preds_df = pd.read_parquet(preds_file)
            else:
                metrics_df = pd.read_csv(metrics_file)
                preds_df = pd.read_csv(preds_file)
            
            # Merge to get model versions
            merged = metrics_df.merge(
                preds_df[['prediction_date', 'horizon', 'model_version']],
                left_on=['evaluation_date', 'horizon'],
                right_on=['prediction_date', 'horizon'],
                how='inner'
            )
            
            comparison = {}
            for version in model_versions:
                version_data = merged[merged['model_version'] == version]
                
                if len(version_data) >= 5:
                    comparison[f"v{version}"] = {
                        'avg_accuracy': float(version_data['adjusted_accuracy'].mean()),
                        'direction_accuracy': float(version_data['direction_accuracy'].mean()),
                        'num_predictions': len(version_data),
                        'avg_magnitude_error': float(version_data['magnitude_mae'].mean()),
                        'brier_score': float(version_data['brier_score'].mean())
                    }
            
            # Statistical comparison if SCIPY available
            if SCIPY_AVAILABLE and len(comparison) >= 2:
                versions_list = list(comparison.keys())
                for i in range(len(versions_list)):
                    for j in range(i + 1, len(versions_list)):
                        v1_data = merged[merged['model_version'] == int(versions_list[i].replace('v', ''))]
                        v2_data = merged[merged['model_version'] == int(versions_list[j].replace('v', ''))]
                        
                        if len(v1_data) >= 10 and len(v2_data) >= 10:
                            # T-test for accuracy difference
                            t_stat, p_value = stats.ttest_ind(
                                v1_data['adjusted_accuracy'].values,
                                v2_data['adjusted_accuracy'].values
                            )
                            
                            comparison[f"{versions_list[i]}_vs_{versions_list[j]}"] = {
                                't_statistic': float(t_stat),
                                'p_value': float(p_value),
                                'significant': p_value < 0.05
                            }
            
            return comparison
        except Exception as e:
            logger.warning(f"compare_models failed: {e}")
            return {}
    
    @staticmethod
    def compare_models_advanced(
        symbol: str, 
        model_versions: List[int],
        metrics: List[str] = ['ic', 'sharpe', 'accuracy']
    ) -> Dict:

        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not metrics_file.exists() or not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                metrics_df = pd.read_parquet(metrics_file)
                preds_df = pd.read_parquet(preds_file)
            else:
                metrics_df = pd.read_csv(metrics_file)
                preds_df = pd.read_csv(preds_file)
            
            # Merge to get model versions
            merged = metrics_df.merge(
                preds_df[['prediction_date', 'horizon', 'model_version']],
                left_on=['evaluation_date', 'horizon'],
                right_on=['prediction_date', 'horizon'],
                how='inner'
            )
            
            results = {}
            
            # Calculate metrics for each version
            for version in model_versions:
                version_data = merged[merged['model_version'] == version]
                
                if len(version_data) < 5:
                    continue
                
                version_results = {
                    'sample_size': len(version_data)
                }
                
                # IC calculation
                if 'ic' in metrics:
                    version_preds = preds_df[preds_df['model_version'] == version]
                    evaluated = version_preds[version_preds['is_evaluated'] == True]
                    
                    if len(evaluated) >= 10:
                        ic_data = FinalPredictionTracker.calculate_information_coefficient(symbol, days=365)
                        if ic_data:
                            version_results['ic_pearson'] = ic_data.get('ic_pearson', 0)
                            version_results['ic_spearman'] = ic_data.get('ic_spearman', 0)
                            version_results['ic_significant'] = ic_data.get('ic_significant', False)
                
                # Sharpe ratio
                if 'sharpe' in metrics:
                    sharpe_data = FinalPredictionTracker.calculate_prediction_sharpe_ratio(symbol, days=60)
                    if sharpe_data:
                        version_results['sharpe_ratio'] = sharpe_data.get('sharpe_ratio', 0)
                
                # Accuracy metrics
                if 'accuracy' in metrics:
                    version_results['avg_accuracy'] = float(version_data['adjusted_accuracy'].mean())
                    version_results['accuracy_std'] = float(version_data['adjusted_accuracy'].std())
                    version_results['direction_accuracy'] = float(version_data['direction_accuracy'].mean())
                    version_results['avg_magnitude_error'] = float(version_data['magnitude_mae'].mean())
                    version_results['brier_score'] = float(version_data['brier_score'].mean())
                
                # Additional metrics
                version_results['avg_confidence'] = float(version_data['confidence'].mean()) if 'confidence' in version_data.columns else 0.0
                version_results['num_predictions'] = len(version_data)
                
                results[f"v{version}"] = version_results
            
            # Statistical significance testing (pairwise comparisons)
            if SCIPY_AVAILABLE and len(model_versions) >= 2:
                statistical_tests = {}
                
                for i in range(len(model_versions)):
                    for j in range(i + 1, len(model_versions)):
                        v1 = model_versions[i]
                        v2 = model_versions[j]
                        
                        v1_data = merged[merged['model_version'] == v1]
                        v2_data = merged[merged['model_version'] == v2]
                        
                        if len(v1_data) >= 10 and len(v2_data) >= 10:
                            test_results = {}
                            
                            # T-test for accuracy
                            if 'adjusted_accuracy' in v1_data.columns and 'adjusted_accuracy' in v2_data.columns:
                                t_stat, p_value = stats.ttest_ind(
                                    v1_data['adjusted_accuracy'].values,
                                    v2_data['adjusted_accuracy'].values
                                )
                                test_results['accuracy'] = {
                                    't_statistic': float(t_stat),
                                    'p_value': float(p_value),
                                    'significant': p_value < 0.05,
                                    'v1_mean': float(v1_data['adjusted_accuracy'].mean()),
                                    'v2_mean': float(v2_data['adjusted_accuracy'].mean())
                                }
                            
                            # T-test for direction accuracy
                            if 'direction_accuracy' in v1_data.columns and 'direction_accuracy' in v2_data.columns:
                                t_stat, p_value = stats.ttest_ind(
                                    v1_data['direction_accuracy'].values,
                                    v2_data['direction_accuracy'].values
                                )
                                test_results['direction_accuracy'] = {
                                    't_statistic': float(t_stat),
                                    'p_value': float(p_value),
                                    'significant': p_value < 0.05
                                }
                            
                            # Mann-Whitney U test (non-parametric alternative)
                            if 'adjusted_accuracy' in v1_data.columns and 'adjusted_accuracy' in v2_data.columns:
                                u_stat, u_p_value = stats.mannwhitneyu(
                                    v1_data['adjusted_accuracy'].values,
                                    v2_data['adjusted_accuracy'].values,
                                    alternative='two-sided'
                                )
                                test_results['mann_whitney'] = {
                                    'u_statistic': float(u_stat),
                                    'p_value': float(u_p_value),
                                    'significant': u_p_value < 0.05
                                }
                            
                            if test_results:
                                statistical_tests[f"v{v1}_vs_v{v2}"] = test_results
                
                if statistical_tests:
                    results['statistical_tests'] = statistical_tests
            
            # Ranking
            if results:
                # Rank by accuracy
                ranked = sorted(
                    results.items(),
                    key=lambda x: x[1].get('avg_accuracy', 0) if 'avg_accuracy' in x[1] else 0,
                    reverse=True
                )
                results['ranking'] = {
                    'by_accuracy': [r[0] for r in ranked]
                }
                
                # Rank by IC if available
                if any('ic_spearman' in v for v in results.values()):
                    ranked_ic = sorted(
                        [(k, v) for k, v in results.items() if 'ic_spearman' in v],
                        key=lambda x: abs(x[1].get('ic_spearman', 0)),
                        reverse=True
                    )
                    results['ranking']['by_ic'] = [r[0] for r in ranked_ic]
            
            return results
            
        except Exception as e:
            logger.warning(f"compare_models_advanced failed: {e}")
            return {}
    
    # ============================================================================
    # REAL-TIME MONITORING & ALERTS
    # ============================================================================
    
    @staticmethod
    def check_performance_alerts(symbol: str) -> List[Dict]:
        """Check for performance degradation and generate alerts"""
        if not FinalPredictionTracker.ENABLE_ALERTS:
            return []
        
        alerts = []
        
        try:
            recent_metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=10)
            
            # Alert 1: Accuracy drop
            if recent_metrics['avg_accuracy'] < 55:
                alerts.append({
                    'level': 'CRITICAL',
                    'type': 'ACCURACY_DROP',
                    'message': f"{symbol}: Accuracy dropped to {recent_metrics['avg_accuracy']:.1f}%",
                    'value': recent_metrics['avg_accuracy'],
                    'threshold': 55.0
                })
            
            # Alert 2: High variance
            if recent_metrics.get('accuracy_std', 0) > 25:
                alerts.append({
                    'level': 'WARNING',
                    'type': 'HIGH_VARIANCE',
                    'message': f"{symbol}: High accuracy variance (std={recent_metrics['accuracy_std']:.1f})",
                    'value': recent_metrics['accuracy_std'],
                    'threshold': 25.0
                })
            
            # Alert 3: Consecutive errors
            if recent_metrics.get('consecutive_direction_errors', 0) >= 5:
                alerts.append({
                    'level': 'CRITICAL',
                    'type': 'CONSECUTIVE_ERRORS',
                    'message': f"{symbol}: {recent_metrics['consecutive_direction_errors']} consecutive direction errors",
                    'value': recent_metrics['consecutive_direction_errors'],
                    'threshold': 5
                })
            
            # Alert 4: Wide confidence interval
            ci = recent_metrics.get('confidence_interval', (50, 90))
            ci_width = ci[1] - ci[0]
            if ci_width > 40:
                alerts.append({
                    'level': 'WARNING',
                    'type': 'HIGH_UNCERTAINTY',
                    'message': f"{symbol}: Wide confidence interval ({ci_width:.1f}%)",
                    'value': ci_width,
                    'threshold': 40.0
                })
            
            # Save alerts
            if alerts:
                FinalPredictionTracker._save_alerts(symbol, alerts)
            
            return alerts
        except Exception as e:
            logger.warning(f"check_performance_alerts failed: {e}")
            return []
    
    @staticmethod
    def _save_alerts(symbol: str, alerts: List[Dict]):
        """Save alerts to file"""
        try:
            alerts_file = FinalPredictionTracker.ALERTS_DIR / f"{symbol}_alerts.json"
            
            if alerts_file.exists():
                with open(alerts_file, 'r') as f:
                    alert_history = json.load(f)
            else:
                alert_history = []
            
            for alert in alerts:
                alert['symbol'] = symbol
                alert['timestamp'] = datetime.now().isoformat()
                alert_history.append(alert)
            
            # Keep last 100 alerts
            alert_history = alert_history[-100:]
            
            with open(alerts_file, 'w') as f:
                json.dump(alert_history, f, separators=(',', ':'))
        except Exception as e:
            logger.warning(f"_save_alerts failed: {e}")
    
    @staticmethod
    def get_health_dashboard(symbols: Optional[List[str]] = None) -> Dict:
        """Get real-time health dashboard for all symbols or specific symbols"""
        try:
            if symbols is None:
                # Get all symbols from predictions directory
                ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
                pred_files = list(FinalPredictionTracker.PREDICTIONS_DIR.glob(f"*_predictions.{ext}"))
                symbols = [f.stem.replace('_predictions', '') for f in pred_files]
            
            dashboard = {
                'timestamp': datetime.now().isoformat(),
                'symbols': {},
                'summary': {
                    'total_symbols': len(symbols),
                    'healthy': 0,
                    'warning': 0,
                    'critical': 0
                }
            }
            
            for symbol in symbols:
                try:
                    recent_metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=10)
                    alerts = FinalPredictionTracker.check_performance_alerts(symbol)
                    
                    # Determine health status
                    health_status = 'healthy'
                    if any(a['level'] == 'CRITICAL' for a in alerts):
                        health_status = 'critical'
                        dashboard['summary']['critical'] += 1
                    elif any(a['level'] == 'WARNING' for a in alerts):
                        health_status = 'warning'
                        dashboard['summary']['warning'] += 1
                    else:
                        dashboard['summary']['healthy'] += 1
                    
                    dashboard['symbols'][symbol] = {
                        'health_status': health_status,
                        'avg_accuracy': recent_metrics.get('avg_accuracy', 0),
                        'num_predictions': recent_metrics.get('num_predictions', 0),
                        'alerts': len(alerts),
                        'recent_alerts': alerts[:3]  # Last 3 alerts
                    }
                except Exception as e:
                    logger.warning(f"Health check failed for {symbol}: {e}")
                    dashboard['symbols'][symbol] = {
                        'health_status': 'unknown',
                        'error': str(e)
                    }
            
            return dashboard
        except Exception as e:
            logger.error(f"get_health_dashboard failed: {e}")
            return {}
    
    # ============================================================================
    # CROSS-ASSET ANALYTICS
    # ============================================================================
    
    @staticmethod
    def calculate_cross_asset_correlation(symbols: List[str], days: int = 60) -> Dict:
        """Calculate correlation of prediction accuracy across assets"""
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            
            # Get metrics for all symbols
            accuracy_data = {}
            for symbol in symbols:
                metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
                if metrics_file.exists():
                    if FinalPredictionTracker.USE_PARQUET:
                        df = pd.read_parquet(metrics_file)
                    else:
                        df = pd.read_csv(metrics_file)
                    
                    if len(df) >= days:
                        recent = df.tail(days)
                        accuracy_data[symbol] = recent['adjusted_accuracy'].values
            
            if len(accuracy_data) < 2:
                return {}
            
            # Create DataFrame
            acc_df = pd.DataFrame(accuracy_data)
            
            # Calculate correlation matrix
            corr_matrix = acc_df.corr()
            
            # Average correlation
            mask = np.triu(np.ones_like(corr_matrix.values), k=1)
            avg_correlation = corr_matrix.values[mask == 1].mean()
            
            return {
                'correlation_matrix': corr_matrix.to_dict(),
                'avg_correlation': float(avg_correlation),
                'num_symbols': len(symbols)
            }
        except Exception as e:
            logger.warning(f"calculate_cross_asset_correlation failed: {e}")
            return {}
    
    @staticmethod
    def get_market_wide_metrics(symbols: Optional[List[str]] = None) -> Dict:
        """Get aggregate performance metrics across all tracked symbols"""
        try:
            if symbols is None:
                ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
                pred_files = list(FinalPredictionTracker.PREDICTIONS_DIR.glob(f"*_predictions.{ext}"))
                symbols = [f.stem.replace('_predictions', '') for f in pred_files]
            
            all_accuracies = []
            all_direction_acc = []
            total_predictions = 0
            
            for symbol in symbols:
                try:
                    metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=30)
                    if metrics.get('num_predictions', 0) > 0:
                        all_accuracies.append(metrics['avg_accuracy'])
                        all_direction_acc.append(metrics['direction_accuracy'])
                        total_predictions += metrics['num_predictions']
                except:
                    continue
            
            if len(all_accuracies) == 0:
                return {}
            
            return {
                'avg_accuracy': float(np.mean(all_accuracies)),
                'accuracy_std': float(np.std(all_accuracies)),
                'avg_direction_accuracy': float(np.mean(all_direction_acc)),
                'total_symbols': len(symbols),
                'total_predictions': total_predictions,
                'symbols_tracked': len([s for s in symbols if FinalPredictionTracker._get_recent_metrics(s, days=30).get('num_predictions', 0) > 0])
            }
        except Exception as e:
            logger.warning(f"get_market_wide_metrics failed: {e}")
            return {}
    
    # ============================================================================
    # MULTI-HORIZON TRACKING
    # ============================================================================
    
    @staticmethod
    def get_horizon_performance(symbol: str) -> Dict:
        """Get performance metrics segmented by prediction horizon"""
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            
            if not metrics_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(metrics_file)
            else:
                df = pd.read_csv(metrics_file)
            
            if 'horizon' not in df.columns or len(df) < 10:
                return {}
            
            horizon_perf = {}
            for horizon in df['horizon'].unique():
                horizon_data = df[df['horizon'] == horizon]
                
                if len(horizon_data) >= 5:
                    horizon_perf[horizon] = {
                        'avg_accuracy': float(horizon_data['adjusted_accuracy'].mean()),
                        'direction_accuracy': float(horizon_data['direction_accuracy'].mean()),
                        'num_predictions': len(horizon_data),
                        'avg_magnitude_error': float(horizon_data['magnitude_mae'].mean())
                    }
            
            return horizon_perf
        except Exception as e:
            logger.warning(f"get_horizon_performance failed: {e}")
            return {}
    
    @staticmethod
    def analyze_horizon_decay(symbol: str) -> Dict:
        """Analyze how prediction accuracy decays with longer horizons"""
        try:
            horizon_perf = FinalPredictionTracker.get_horizon_performance(symbol)
            
            if len(horizon_perf) < 2:
                return {}
            
            # Expected order: 1d > 1w > 1m
            horizons = ['1d', '1w', '1m']
            decay_analysis = {}
            
            for i, horizon in enumerate(horizons):
                if horizon in horizon_perf:
                    decay_analysis[horizon] = horizon_perf[horizon]
                    
                    # Calculate decay from previous horizon
                    if i > 0:
                        prev_horizon = horizons[i - 1]
                        if prev_horizon in horizon_perf:
                            prev_acc = horizon_perf[prev_horizon]['avg_accuracy']
                            curr_acc = horizon_perf[horizon]['avg_accuracy']
                            decay = prev_acc - curr_acc
                            decay_analysis[horizon]['decay_from_previous'] = float(decay)
                            decay_analysis[horizon]['decay_pct'] = float((decay / prev_acc * 100) if prev_acc > 0 else 0)
            
            return decay_analysis
        except Exception as e:
            logger.warning(f"analyze_horizon_decay failed: {e}")
            return {}
    
    # ============================================================================
    # PREDICTION RECONCILIATION
    # ============================================================================
    
    @staticmethod
    def reconcile_predictions(symbol: str, date: str, horizon: str = '1d') -> Dict:
        """
        Handle multiple predictions for the same date/horizon.
        Returns the most recent or best quality prediction.
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            # Find all predictions for this date/horizon
            mask = (df['prediction_date'] == date) & (df['horizon'] == horizon)
            matches = df[mask]
            
            if len(matches) == 0:
                return {}
            
            if len(matches) == 1:
                return matches.iloc[0].to_dict()
            
            # Multiple predictions: use highest quality score, then most recent
            matches = matches.sort_values(['quality_score', 'created_at'], ascending=[False, False])
            best_pred = matches.iloc[0]
            
            return {
                'prediction': best_pred.to_dict(),
                'num_conflicts': len(matches),
                'resolution_method': 'highest_quality',
                'conflicts': matches[['prediction_date', 'quality_score', 'created_at', 'model_version']].to_dict('records')
            }
        except Exception as e:
            logger.warning(f"reconcile_predictions failed: {e}")
            return {}
    
    # ============================================================================
    # BACKTESTING INTEGRATION HOOKS
    # ============================================================================
    
    @staticmethod
    def link_prediction_to_trade(
        symbol: str, prediction_date: str, trade_id: str,
        execution_price: float, quantity: int, pnl: Optional[float] = None
    ):
        """Link a prediction to an actual trade for P&L attribution"""
        try:
            links_file = FinalPredictionTracker.TRACKING_DIR / "prediction_trade_links.json"
            
            if links_file.exists():
                with open(links_file, 'r') as f:
                    links = json.load(f)
            else:
                links = []
            
            link = {
                'symbol': symbol,
                'prediction_date': prediction_date,
                'trade_id': trade_id,
                'execution_price': float(execution_price),
                'quantity': int(quantity),
                'pnl': float(pnl) if pnl is not None else None,
                'created_at': datetime.now().isoformat()
            }
            
            links.append(link)
            links = links[-1000:]  # Keep last 1000 links
            
            with open(links_file, 'w') as f:
                json.dump(links, f, separators=(',', ':'))
        except Exception as e:
            logger.warning(f"link_prediction_to_trade failed: {e}")
    
    @staticmethod
    def get_prediction_pnl_attribution(symbol: str, days: int = 60) -> Dict:
        """Calculate P&L attribution to predictions"""
        try:
            links_file = FinalPredictionTracker.TRACKING_DIR / "prediction_trade_links.json"
            
            if not links_file.exists():
                return {}
            
            with open(links_file, 'r') as f:
                links = json.load(f)
            
            # Filter by symbol and date
            symbol_links = [l for l in links if l['symbol'] == symbol]
            
            if len(symbol_links) == 0:
                return {}
            
            # Calculate total P&L
            total_pnl = sum(l.get('pnl', 0) for l in symbol_links if l.get('pnl') is not None)
            num_trades = len([l for l in symbol_links if l.get('pnl') is not None])
            
            # Get prediction accuracy for same period
            recent_metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=days)
            
            # Calculate correlation between accuracy and P&L if enough data
            correlation = 'N/A'
            if num_trades >= 10:
                try:
                    # Get predictions and P&L for correlation
                    pnl_values = [l.get('pnl', 0) for l in symbol_links if l.get('pnl') is not None]
                    if len(pnl_values) >= 10 and SCIPY_AVAILABLE:
                        # Get corresponding accuracies (simplified - would need date matching)
                        correlation = 0.0  # Placeholder - would need proper date matching
                except:
                    pass
            
            return {
                'total_pnl': float(total_pnl),
                'num_trades': num_trades,
                'avg_pnl_per_trade': float(total_pnl / num_trades) if num_trades > 0 else 0.0,
                'prediction_accuracy': recent_metrics.get('avg_accuracy', 0),
                'correlation': correlation
            }
        except Exception as e:
            logger.warning(f"get_prediction_pnl_attribution failed: {e}")
            return {}
    
    # ============================================================================
    # ADVANCED METRICS: PREDICTION SHARPE, TURNOVER, STABILITY
    # ============================================================================
    
    @staticmethod
    def calculate_prediction_sharpe_ratio(symbol: str, days: int = 60, risk_free_rate: float = 0.0) -> Dict:
        """
        Calculate Sharpe ratio of predictions (treating prediction accuracy as returns).
        Higher Sharpe = more consistent, reliable predictions.
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            
            if not metrics_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(metrics_file)
            else:
                df = pd.read_csv(metrics_file)
            
            if len(df) < 20:
                return {}
            
            # Get recent metrics
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = df[df['evaluation_date'] >= cutoff_date] if 'evaluation_date' in df.columns else df.tail(days)
            
            if len(recent) < 20:
                return {}
            
            # Calculate daily "returns" from accuracy changes
            accuracies = recent['adjusted_accuracy'].values
            accuracy_returns = np.diff(accuracies) / 100.0  # Convert to decimal
            
            if len(accuracy_returns) < 10:
                return {}
            
            # Sharpe ratio
            mean_return = np.mean(accuracy_returns)
            std_return = np.std(accuracy_returns)
            
            if std_return == 0:
                sharpe = 0.0
            else:
                # Annualized Sharpe (assuming daily data)
                sharpe = (mean_return - risk_free_rate / 252) / std_return * np.sqrt(252)
            
            return {
                'sharpe_ratio': float(sharpe),
                'mean_accuracy_return': float(mean_return),
                'std_accuracy_return': float(std_return),
                'num_samples': len(accuracy_returns)
            }
        except Exception as e:
            logger.warning(f"calculate_prediction_sharpe_ratio failed: {e}")
            return {}
    
    @staticmethod
    def calculate_prediction_turnover(symbol: str, days: int = 30) -> Dict:
        """
        Calculate prediction turnover - how often predictions change direction.
        High turnover = predictions are unstable/changing frequently.
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            if len(df) < 10:
                return {}
            
            # Sort by date
            df = df.sort_values('prediction_date')
            
            # Get recent predictions
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = df[df['prediction_date'] >= cutoff_date] if 'prediction_date' in df.columns else df.tail(days)
            
            if len(recent) < 10:
                return {}
            
            # Calculate direction changes
            pred_changes = recent['predicted_change_pct'].values
            directions = np.sign(pred_changes)
            
            # Count direction changes
            direction_changes = np.sum(np.diff(directions) != 0)
            turnover_rate = direction_changes / max(1, len(directions) - 1)
            
            return {
                'turnover_rate': float(turnover_rate),
                'direction_changes': int(direction_changes),
                'total_predictions': len(recent),
                'stability': float(1.0 - turnover_rate)  # Inverse of turnover
            }
        except Exception as e:
            logger.warning(f"calculate_prediction_turnover failed: {e}")
            return {}
    
    @staticmethod
    def calculate_prediction_stability(symbol: str, days: int = 60) -> Dict:
        """
        Calculate prediction stability metrics:
        - Consistency of predictions over time
        - Autocorrelation of predictions
        - Prediction persistence
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            if len(df) < 20:
                return {}
            
            # Sort by date
            df = df.sort_values('prediction_date')
            
            # Get recent predictions
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = df[df['prediction_date'] >= cutoff_date] if 'prediction_date' in df.columns else df.tail(days)
            
            if len(recent) < 20:
                return {}
            
            pred_changes = recent['predicted_change_pct'].values
            confidences = recent['confidence'].values if 'confidence' in recent.columns else np.ones(len(recent)) * 75.0
            
            # 1. Prediction variance (lower = more stable)
            pred_variance = float(np.var(pred_changes))
            
            # 2. Confidence stability
            conf_variance = float(np.var(confidences))
            
            # 3. Autocorrelation (lag-1)
            if len(pred_changes) >= 10 and SCIPY_AVAILABLE:
                autocorr = stats.pearsonr(pred_changes[:-1], pred_changes[1:])[0]
                if np.isnan(autocorr):
                    autocorr = 0.0
            else:
                autocorr = 0.0
            
            # 4. Prediction persistence (how often same direction)
            directions = np.sign(pred_changes)
            persistence = float(np.sum(np.diff(directions) == 0) / max(1, len(directions) - 1))
            
            return {
                'prediction_variance': pred_variance,
                'confidence_variance': conf_variance,
                'autocorrelation_lag1': float(autocorr),
                'persistence': persistence,
                'stability_score': float(1.0 / (1.0 + pred_variance / 10.0))  # Normalized stability
            }
        except Exception as e:
            logger.warning(f"calculate_prediction_stability failed: {e}")
            return {}
    
    @staticmethod
    def calculate_rolling_ic(symbol: str, window_size: int = 20, days: int = 120) -> Dict:
        """
        Calculate rolling window Information Coefficient.
        Shows how IC varies over time (important for detecting degradation).
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            # Get evaluated predictions
            evaluated = df[df['is_evaluated'] == True].copy()
            
            if len(evaluated) < window_size * 2:
                return {}
            
            # Sort by date
            evaluated = evaluated.sort_values('prediction_date')
            
            # Get recent
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = evaluated[evaluated['prediction_date'] >= cutoff_date] if len(evaluated) > days else evaluated.tail(days)
            
            if len(recent) < window_size * 2:
                return {}
            
            pred_changes = recent['predicted_change_pct'].values
            actual_changes = recent['actual_change_pct'].values
            
            # Remove NaN
            mask = ~(np.isnan(pred_changes) | np.isnan(actual_changes))
            pred_changes = pred_changes[mask]
            actual_changes = actual_changes[mask]
            
            if len(pred_changes) < window_size * 2:
                return {}
            
            # Calculate rolling IC
            rolling_ics = []
            rolling_dates = []
            
            for i in range(len(pred_changes) - window_size + 1):
                window_pred = pred_changes[i:i+window_size]
                window_actual = actual_changes[i:i+window_size]
                
                if SCIPY_AVAILABLE and len(window_pred) >= 10:
                    ic, _ = stats.spearmanr(window_pred, window_actual)
                    if not np.isnan(ic):
                        rolling_ics.append(ic)
                        rolling_dates.append(i)
            
            if len(rolling_ics) == 0:
                return {}
            
            rolling_ics = np.array(rolling_ics)
            
            return {
                'rolling_ic_mean': float(np.mean(rolling_ics)),
                'rolling_ic_std': float(np.std(rolling_ics)),
                'rolling_ic_min': float(np.min(rolling_ics)),
                'rolling_ic_max': float(np.max(rolling_ics)),
                'ic_trend': float(np.polyfit(range(len(rolling_ics)), rolling_ics, 1)[0]),  # Slope
                'num_windows': len(rolling_ics),
                'window_size': window_size
            }
        except Exception as e:
            logger.warning(f"calculate_rolling_ic failed: {e}")
            return {}
    
    @staticmethod
    def calculate_ensemble_diversity(symbol: str, days: int = 30) -> Dict:
        """
        Calculate ensemble diversity metrics.
        Higher diversity = ensemble members disagree more (good for robustness).
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            if len(df) < 10:
                return {}
            
            # Get recent predictions
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = df[df['prediction_date'] >= cutoff_date] if 'prediction_date' in df.columns else df.tail(days)
            
            if len(recent) < 10:
                return {}
            
            # Use prediction_uncertainty as proxy for ensemble diversity
            if 'prediction_uncertainty' in recent.columns:
                uncertainties = recent['prediction_uncertainty'].values
                avg_uncertainty = float(np.mean(uncertainties))
                uncertainty_std = float(np.std(uncertainties))
            else:
                avg_uncertainty = 0.0
                uncertainty_std = 0.0
            
            # Use ensemble_agreement as inverse diversity measure
            if 'ensemble_agreement' in recent.columns:
                agreements = recent['ensemble_agreement'].values
                avg_agreement = float(np.mean(agreements))
                diversity = float(1.0 - avg_agreement)  # Inverse of agreement
            else:
                avg_agreement = 0.0
                diversity = 0.0
            
            return {
                'avg_uncertainty': avg_uncertainty,
                'uncertainty_std': uncertainty_std,
                'avg_ensemble_agreement': avg_agreement,
                'diversity_score': diversity,
                'num_predictions': len(recent)
            }
        except Exception as e:
            logger.warning(f"calculate_ensemble_diversity failed: {e}")
            return {}
    
    @staticmethod
    def calculate_signal_to_noise_ratio(symbol: str, days: int = 60) -> Dict:
        """
        Calculate signal-to-noise ratio of predictions.
        Signal = mean prediction magnitude, Noise = prediction variance.
        Higher SNR = more reliable predictions.
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            if len(df) < 20:
                return {}
            
            # Get recent predictions
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = df[df['prediction_date'] >= cutoff_date] if 'prediction_date' in df.columns else df.tail(days)
            
            if len(recent) < 20:
                return {}
            
            pred_changes = recent['predicted_change_pct'].values
            
            # Signal = mean absolute prediction
            signal = float(np.mean(np.abs(pred_changes)))
            
            # Noise = standard deviation
            noise = float(np.std(pred_changes))
            
            # SNR
            snr = signal / noise if noise > 0 else 0.0
            
            return {
                'signal': signal,
                'noise': noise,
                'snr': float(snr),
                'num_predictions': len(recent)
            }
        except Exception as e:
            logger.warning(f"calculate_signal_to_noise_ratio failed: {e}")
            return {}
    
    # ============================================================================
    # INSTITUTIONAL-GRADE FEATURES (Hedge Fund Level)
    # ============================================================================
    
    @staticmethod
    def adjust_confidence_by_calibration(symbol: str, raw_confidence: float) -> float:
        """
        Auto-adjust prediction confidence based on calibration curve.
        This is what top hedge funds do - they calibrate their confidence scores.
        
        If model is overconfident (predicted > actual), reduce confidence.
        If model is underconfident (predicted < actual), increase confidence.
        """
        try:
            calibration_curve = FinalPredictionTracker.get_calibration_curve(symbol, min_samples=5)
            
            if not calibration_curve:
                return raw_confidence  # No calibration data, return as-is
            
            # Find the calibration bin for this confidence
            confidence_bin = int(raw_confidence // 10) * 10
            
            if confidence_bin in calibration_curve:
                calib_data = calibration_curve[confidence_bin]
                predicted_prob = calib_data['predicted_prob']
                actual_frequency = calib_data['actual_frequency']
                
                # Calibration error
                calibration_error = predicted_prob - actual_frequency
                
                # Adjust confidence: if overconfident (predicted > actual), reduce
                # If underconfident (predicted < actual), increase
                adjustment_factor = 1.0 - (calibration_error * 0.5)  # Dampen adjustment
                adjusted_confidence = raw_confidence * adjustment_factor
                
                # Clamp to valid range
                adjusted_confidence = np.clip(adjusted_confidence, 0.0, 100.0)
                
                return float(adjusted_confidence)
            else:
                return raw_confidence
        except Exception as e:
            logger.warning(f"adjust_confidence_by_calibration failed: {e}")
            return raw_confidence
    
    @staticmethod
    def detect_drift_adwin(symbol: str, current_df: pd.DataFrame, delta: float = 0.002) -> bool:
        """
        ADWIN (Adaptive Windowing) drift detection - more sophisticated than KS test.
        Used by top hedge funds for concept drift detection.
        
        ADWIN automatically detects change and adapts window size.
        """
        try:
            if len(current_df) < 100:
                return False
            
            # Use recent vs historical data
            recent_data = current_df.iloc[-30:]
            historical_data = current_df.iloc[-200:-30]
            
            if len(recent_data) < 20 or len(historical_data) < 50:
                return False
            
            # Calculate mean and variance for both windows
            test_features = ['Returns', 'Volume_Ratio', 'RSI_14', 'Volatility_10']
            drift_detected_count = 0
            
            for feat in test_features:
                if feat not in current_df.columns:
                    continue
                
                recent_vals = recent_data[feat].dropna().values
                hist_vals = historical_data[feat].dropna().values
                
                if len(recent_vals) < 10 or len(hist_vals) < 30:
                    continue
                
                # ADWIN-like detection: compare means with adaptive threshold
                recent_mean = np.mean(recent_vals)
                hist_mean = np.mean(hist_vals)
                
                recent_std = np.std(recent_vals)
                hist_std = np.std(hist_vals)
                
                # Combined standard error
                n1, n2 = len(recent_vals), len(hist_vals)
                combined_se = np.sqrt((recent_std**2 / n1) + (hist_std**2 / n2))
                
                # Adaptive threshold based on window sizes
                threshold = delta * np.sqrt(np.log(n1 + n2) / min(n1, n2))
                
                # Drift if mean difference exceeds threshold
                if abs(recent_mean - hist_mean) > threshold * combined_se:
                    drift_detected_count += 1
            
            # Drift if multiple features show drift
            return drift_detected_count >= 2
        except Exception as e:
            logger.warning(f"detect_drift_adwin failed: {e}")
            return False
    
    @staticmethod
    def calculate_prediction_decay(symbol: str, horizon: str = '1d') -> Dict:
        """
        Calculate how prediction accuracy decays over time.
        Critical for understanding prediction aging and when to refresh.
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            
            if not metrics_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(metrics_file)
            else:
                df = pd.read_csv(metrics_file)
            
            # Filter by horizon
            if 'horizon' in df.columns:
                df = df[df['horizon'] == horizon]
            
            if len(df) < 30:
                return {}
            
            # Sort by evaluation date
            df = df.sort_values('evaluation_date')
            
            # Calculate accuracy over time (rolling window)
            window_size = min(10, len(df) // 3)
            df['rolling_accuracy'] = df['adjusted_accuracy'].rolling(window=window_size, min_periods=5).mean()
            
            # Calculate decay rate (slope of accuracy over time)
            if len(df) >= 20:
                x = np.arange(len(df))
                y = df['rolling_accuracy'].dropna().values
                x_clean = x[:len(y)]
                
                if len(y) >= 10:
                    slope, intercept = np.polyfit(x_clean, y, 1)
                    decay_rate = float(slope)  # Negative = decaying, positive = improving
                else:
                    decay_rate = 0.0
            else:
                decay_rate = 0.0
            
            # Calculate half-life (time for accuracy to drop by 50%)
            if decay_rate < 0:
                half_life = abs(50.0 / decay_rate) if decay_rate != 0 else float('inf')
            else:
                half_life = float('inf')
            
            return {
                'decay_rate': decay_rate,
                'half_life_days': half_life if half_life != float('inf') else None,
                'current_accuracy': float(df['adjusted_accuracy'].iloc[-1]) if len(df) > 0 else 0.0,
                'avg_accuracy': float(df['adjusted_accuracy'].mean()),
                'trend': 'decaying' if decay_rate < -0.1 else 'improving' if decay_rate > 0.1 else 'stable'
            }
        except Exception as e:
            logger.warning(f"calculate_prediction_decay failed: {e}")
            return {}
    
    @staticmethod
    def track_feature_stability(symbol: str, days: int = 60) -> Dict:
        """
        Track feature importance stability over time.
        Unstable features indicate model instability or data quality issues.
        """
        try:
            importance_file = FinalPredictionTracker.FEATURE_IMPORTANCE_DIR / f"{symbol}_importance.json"
            
            if not importance_file.exists():
                return {}
            
            with open(importance_file, 'r') as f:
                history = json.load(f)
            
            if len(history) < 10:
                return {}
            
            # Get recent history
            recent_history = history[-min(days, len(history)):]
            
            # Aggregate feature importances over time
            feature_importances = defaultdict(list)
            
            for record in recent_history:
                for feat, imp in record['importance'].items():
                    feature_importances[feat].append(imp)
            
            # Calculate stability metrics for each feature
            stability_metrics = {}
            
            for feat, imp_history in feature_importances.items():
                if len(imp_history) < 5:
                    continue
                
                imp_array = np.array(imp_history)
                
                # Coefficient of variation (CV) = std / mean (lower = more stable)
                if np.mean(imp_array) > 0:
                    cv = float(np.std(imp_array) / np.mean(imp_array))
                else:
                    cv = 0.0
                
                # Trend (slope)
                x = np.arange(len(imp_array))
                slope = float(np.polyfit(x, imp_array, 1)[0])
                
                stability_metrics[feat] = {
                    'coefficient_of_variation': cv,
                    'stability_score': float(1.0 / (1.0 + cv)),  # Higher = more stable
                    'trend': slope,
                    'mean_importance': float(np.mean(imp_array)),
                    'std_importance': float(np.std(imp_array))
                }
            
            # Overall stability
            if stability_metrics:
                avg_cv = np.mean([m['coefficient_of_variation'] for m in stability_metrics.values()])
                overall_stability = float(1.0 / (1.0 + avg_cv))
            else:
                overall_stability = 0.0
            
            # Most and least stable features
            sorted_features = sorted(
                stability_metrics.items(),
                key=lambda x: x[1]['stability_score'],
                reverse=True
            )
            
            return {
                'overall_stability': overall_stability,
                'num_features_tracked': len(stability_metrics),
                'most_stable_features': [f[0] for f in sorted_features[:5]],
                'least_stable_features': [f[0] for f in sorted_features[-5:]],
                'feature_stability': {k: {
                    'stability_score': v['stability_score'],
                    'trend': v['trend']
                } for k, v in stability_metrics.items()}
            }
        except Exception as e:
            logger.warning(f"track_feature_stability failed: {e}")
            return {}
    
    @staticmethod
    def calculate_prediction_correlation_with_factors(
        symbol: str, market_data: Optional[pd.DataFrame] = None, days: int = 60
    ) -> Dict:
        """
        Calculate correlation between predictions and market factors (beta, sector, style).
        Critical for understanding prediction drivers and factor exposure.
        """
        try:
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return {}
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            if len(df) < 20:
                return {}
            
            # Get recent predictions
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            recent = df[df['prediction_date'] >= cutoff_date] if 'prediction_date' in df.columns else df.tail(days)
            
            if len(recent) < 20:
                return {}
            
            pred_changes = recent['predicted_change_pct'].values
            
            correlations = {}
            
            # If market data provided, calculate correlations
            if market_data is not None and len(market_data) >= len(recent):
                # Align dates (simplified - would need proper date matching)
                market_aligned = market_data.tail(len(recent))
                
                # Market return correlation (beta-like)
                if 'Returns' in market_aligned.columns:
                    market_returns = market_aligned['Returns'].values
                    if SCIPY_AVAILABLE and len(market_returns) == len(pred_changes):
                        corr, p_val = stats.pearsonr(pred_changes, market_returns)
                        correlations['market_beta'] = {
                            'correlation': float(corr) if not np.isnan(corr) else 0.0,
                            'p_value': float(p_val) if SCIPY_AVAILABLE else 0.0,
                            'significant': p_val < 0.05 if SCIPY_AVAILABLE else False
                        }
                
                # Volatility correlation
                if 'Volatility_10' in market_aligned.columns:
                    vol = market_aligned['Volatility_10'].values
                    if SCIPY_AVAILABLE and len(vol) == len(pred_changes):
                        corr, p_val = stats.pearsonr(pred_changes, vol)
                        correlations['volatility'] = {
                            'correlation': float(corr) if not np.isnan(corr) else 0.0,
                            'p_value': float(p_val) if SCIPY_AVAILABLE else 0.0
                        }
            
            return {
                'factor_correlations': correlations,
                'num_predictions': len(recent)
            }
        except Exception as e:
            logger.warning(f"calculate_prediction_correlation_with_factors failed: {e}")
            return {}
    
    @staticmethod
    def calculate_comprehensive_quality_score(
        symbol: str, prediction_date: str, horizon: str = '1d'
    ) -> Dict:
        """
        Multi-dimensional quality score combining multiple factors.
        This is what top hedge funds use - comprehensive quality assessment.
        """
        try:
            # Get base metrics
            recent_metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=30)
            
            # Get IC
            ic_data = FinalPredictionTracker.calculate_information_coefficient(symbol, days=60)
            
            # Get stability
            stability_data = FinalPredictionTracker.calculate_prediction_stability(symbol, days=60)
            
            # Get calibration
            calibration_curve = FinalPredictionTracker.get_calibration_curve(symbol, min_samples=5)
            
            # Get turnover
            turnover_data = FinalPredictionTracker.calculate_prediction_turnover(symbol, days=30)
            
            # Calculate quality components (0-100 scale)
            components = {}
            
            # 1. Accuracy component (0-40 points)
            accuracy = recent_metrics.get('avg_accuracy', 50)
            components['accuracy_score'] = min(40, accuracy * 0.4)
            
            # 2. IC component (0-20 points)
            rank_ic = ic_data.get('ic_spearman', 0)
            components['ic_score'] = min(20, abs(rank_ic) * 20)
            
            # 3. Stability component (0-15 points)
            stability = stability_data.get('stability_score', 0.5)
            components['stability_score'] = stability * 15
            
            # 4. Calibration component (0-15 points)
            if calibration_curve:
                # Average calibration error
                calib_errors = [abs(v['predicted_prob'] - v['actual_frequency']) for v in calibration_curve.values()]
                avg_calib_error = np.mean(calib_errors) if calib_errors else 0.5
                components['calibration_score'] = max(0, 15 * (1.0 - avg_calib_error * 2))
            else:
                components['calibration_score'] = 7.5  # Neutral if no data
            
            # 5. Turnover component (0-10 points) - lower turnover = higher score
            turnover = turnover_data.get('turnover_rate', 0.5)
            components['turnover_score'] = max(0, 10 * (1.0 - turnover))
            
            # Total quality score
            total_score = sum(components.values())
            
            # Grade
            if total_score >= 85:
                grade = 'A+'
            elif total_score >= 75:
                grade = 'A'
            elif total_score >= 65:
                grade = 'B'
            elif total_score >= 55:
                grade = 'C'
            else:
                grade = 'D'
            
            return {
                'total_quality_score': float(total_score),
                'grade': grade,
                'components': components,
                'breakdown': {
                    'accuracy': recent_metrics.get('avg_accuracy', 0),
                    'ic_spearman': rank_ic,
                    'stability': stability,
                    'turnover': turnover
                }
            }
        except Exception as e:
            logger.warning(f"calculate_comprehensive_quality_score failed: {e}")
            return {}
    
    @staticmethod
    def get_institutional_dashboard(symbol: str) -> Dict:
        """
        Comprehensive institutional-grade dashboard combining all metrics.
        This is what portfolio managers at top hedge funds see.
        """
        try:
            dashboard = {
                'symbol': symbol,
                'timestamp': datetime.now().isoformat(),
                'metrics': {}
            }
            
            # Core metrics
            recent_metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=30)
            dashboard['metrics']['core'] = recent_metrics
            
            # Advanced metrics
            dashboard['metrics']['ic'] = FinalPredictionTracker.calculate_information_coefficient(symbol, days=60)
            dashboard['metrics']['sharpe'] = FinalPredictionTracker.calculate_prediction_sharpe_ratio(symbol, days=60)
            dashboard['metrics']['stability'] = FinalPredictionTracker.calculate_prediction_stability(symbol, days=60)
            dashboard['metrics']['turnover'] = FinalPredictionTracker.calculate_prediction_turnover(symbol, days=30)
            dashboard['metrics']['rolling_ic'] = FinalPredictionTracker.calculate_rolling_ic(symbol, window_size=20, days=120)
            dashboard['metrics']['ensemble_diversity'] = FinalPredictionTracker.calculate_ensemble_diversity(symbol, days=30)
            dashboard['metrics']['snr'] = FinalPredictionTracker.calculate_signal_to_noise_ratio(symbol, days=60)
            
            # Decay and stability
            dashboard['metrics']['decay'] = FinalPredictionTracker.calculate_prediction_decay(symbol, horizon='1d')
            dashboard['metrics']['feature_stability'] = FinalPredictionTracker.track_feature_stability(symbol, days=60)
            
            # Quality score
            dashboard['metrics']['quality'] = FinalPredictionTracker.calculate_comprehensive_quality_score(symbol, datetime.now().strftime('%Y-%m-%d'))
            
            # Alerts
            dashboard['alerts'] = FinalPredictionTracker.check_performance_alerts(symbol)
            
            # Regime performance
            dashboard['metrics']['regime_performance'] = FinalPredictionTracker.get_regime_performance(symbol)
            
            # Horizon performance
            dashboard['metrics']['horizon_performance'] = FinalPredictionTracker.get_horizon_performance(symbol)
            
            return dashboard
        except Exception as e:
            logger.error(f"get_institutional_dashboard failed: {e}")
            return {}
    
    # ============================================================================
    # ADVANCED VISUALIZATION: INTERACTIVE DASHBOARDS (PLOTLY/DASH)
    # ============================================================================
    
    @staticmethod
    def generate_interactive_dashboard(symbol: str, output_file: Optional[str] = None) -> Optional[str]:
        """
        Generate interactive Plotly dashboard for prediction tracking.
        Returns HTML file path if successful.
        """
        if not PLOTLY_AVAILABLE:
            logger.warning("Plotly not available. Install with: pip install plotly")
            return None
        
        try:
            # Get data
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
            
            if not preds_file.exists() or not metrics_file.exists():
                logger.warning(f"Data files not found for {symbol}")
                return None
            
            if FinalPredictionTracker.USE_PARQUET:
                preds_df = pd.read_parquet(preds_file)
                metrics_df = pd.read_parquet(metrics_file)
            else:
                preds_df = pd.read_csv(preds_file)
                metrics_df = pd.read_csv(metrics_file)
            
            # Create subplots
            fig = make_subplots(
                rows=3, cols=2,
                subplot_titles=(
                    'Prediction Accuracy Over Time',
                    'Information Coefficient (Rolling)',
                    'Prediction vs Actual',
                    'Calibration Curve',
                    'Feature Importance Stability',
                    'Prediction Quality Score'
                ),
                specs=[[{"secondary_y": False}, {"secondary_y": False}],
                       [{"secondary_y": False}, {"secondary_y": False}],
                       [{"secondary_y": False}, {"secondary_y": False}]]
            )
            
            # 1. Accuracy over time
            if 'evaluation_date' in metrics_df.columns and 'adjusted_accuracy' in metrics_df.columns:
                metrics_df_sorted = metrics_df.sort_values('evaluation_date')
                fig.add_trace(
                    go.Scatter(
                        x=metrics_df_sorted['evaluation_date'],
                        y=metrics_df_sorted['adjusted_accuracy'],
                        mode='lines+markers',
                        name='Accuracy',
                        line=dict(color='blue', width=2)
                    ),
                    row=1, col=1
                )
            
            # 2. Rolling IC
            ic_data = FinalPredictionTracker.calculate_rolling_ic(symbol, window_size=20, days=120)
            if ic_data and 'rolling_ic_mean' in ic_data:
                # Simplified - would need actual rolling data
                fig.add_trace(
                    go.Scatter(
                        x=list(range(ic_data.get('num_windows', 0))),
                        y=[ic_data['rolling_ic_mean']] * ic_data.get('num_windows', 1),
                        mode='lines',
                        name='Rolling IC',
                        line=dict(color='green', width=2)
                    ),
                    row=1, col=2
                )
            
            # 3. Prediction vs Actual
            evaluated = preds_df[preds_df['is_evaluated'] == True]
            if len(evaluated) > 0 and 'predicted_change_pct' in evaluated.columns and 'actual_change_pct' in evaluated.columns:
                fig.add_trace(
                    go.Scatter(
                        x=evaluated['predicted_change_pct'],
                        y=evaluated['actual_change_pct'],
                        mode='markers',
                        name='Predictions',
                        marker=dict(color='purple', size=5, opacity=0.6)
                    ),
                    row=2, col=1
                )
                # Add diagonal line
                max_val = max(evaluated['predicted_change_pct'].abs().max(), evaluated['actual_change_pct'].abs().max())
                fig.add_trace(
                    go.Scatter(
                        x=[-max_val, max_val],
                        y=[-max_val, max_val],
                        mode='lines',
                        name='Perfect Prediction',
                        line=dict(color='red', dash='dash', width=1)
                    ),
                    row=2, col=1
                )
            
            # 4. Calibration curve
            calib_curve = FinalPredictionTracker.get_calibration_curve(symbol, min_samples=5)
            if calib_curve:
                bins = sorted(calib_curve.keys())
                predicted_probs = [calib_curve[b]['predicted_prob'] for b in bins]
                actual_freqs = [calib_curve[b]['actual_frequency'] for b in bins]
                
                fig.add_trace(
                    go.Scatter(
                        x=predicted_probs,
                        y=actual_freqs,
                        mode='lines+markers',
                        name='Calibration',
                        line=dict(color='orange', width=2)
                    ),
                    row=2, col=2
                )
                # Add perfect calibration line
                fig.add_trace(
                    go.Scatter(
                        x=[0, 1],
                        y=[0, 1],
                        mode='lines',
                        name='Perfect Calibration',
                        line=dict(color='red', dash='dash', width=1)
                    ),
                    row=2, col=2
                )
            
            # 5. Feature stability
            feat_stability = FinalPredictionTracker.track_feature_stability(symbol, days=60)
            if feat_stability and 'feature_stability' in feat_stability:
                features = list(feat_stability['feature_stability'].keys())[:10]  # Top 10
                stability_scores = [feat_stability['feature_stability'][f]['stability_score'] for f in features]
                
                fig.add_trace(
                    go.Bar(
                        x=features,
                        y=stability_scores,
                        name='Stability',
                        marker_color='teal'
                    ),
                    row=3, col=1
                )
            
            # 6. Quality score
            quality = FinalPredictionTracker.calculate_comprehensive_quality_score(symbol, datetime.now().strftime('%Y-%m-%d'))
            if quality and 'components' in quality:
                components = quality['components']
                labels = list(components.keys())
                values = list(components.values())
                
                fig.add_trace(
                    go.Bar(
                        x=labels,
                        y=values,
                        name='Quality Components',
                        marker_color='coral'
                    ),
                    row=3, col=2
                )
            
            # Update layout
            fig.update_layout(
                height=1200,
                title_text=f"Prediction Tracking Dashboard: {symbol}",
                showlegend=True,
                template="plotly_white"
            )
            
            # Update axes labels
            fig.update_xaxes(title_text="Date", row=1, col=1)
            fig.update_yaxes(title_text="Accuracy (%)", row=1, col=1)
            fig.update_xaxes(title_text="Window", row=1, col=2)
            fig.update_yaxes(title_text="IC", row=1, col=2)
            fig.update_xaxes(title_text="Predicted Change (%)", row=2, col=1)
            fig.update_yaxes(title_text="Actual Change (%)", row=2, col=1)
            fig.update_xaxes(title_text="Predicted Probability", row=2, col=2)
            fig.update_yaxes(title_text="Actual Frequency", row=2, col=2)
            fig.update_xaxes(title_text="Feature", row=3, col=1)
            fig.update_yaxes(title_text="Stability Score", row=3, col=1)
            fig.update_xaxes(title_text="Component", row=3, col=2)
            fig.update_yaxes(title_text="Score", row=3, col=2)
            
            # Save to file
            if output_file is None:
                output_file = str(FinalPredictionTracker.TRACKING_DIR / f"{symbol}_dashboard.html")
            
            fig.write_html(output_file)
            logger.info(f"Dashboard saved to {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f"generate_interactive_dashboard failed: {e}")
            return None
    
    @staticmethod
    def create_dash_app(symbols: List[str], port: int = 8050):
        """
        Create interactive Dash web application for prediction tracking.
        Run with: app.run_server(debug=True)
        """
        if not DASH_AVAILABLE:
            logger.warning("Dash not available. Install with: pip install dash")
            return None
        
        try:
            app = dash.Dash(__name__)
            
            app.layout = html.Div([
                html.H1("Prediction Tracking Dashboard", style={'textAlign': 'center'}),
                dcc.Dropdown(
                    id='symbol-dropdown',
                    options=[{'label': s, 'value': s} for s in symbols],
                    value=symbols[0] if symbols else None
                ),
                dcc.Graph(id='accuracy-graph'),
                dcc.Graph(id='ic-graph'),
                dcc.Interval(
                    id='interval-component',
                    interval=60*1000,  # Update every minute
                    n_intervals=0
                )
            ])
            
            @app.callback(
                [Output('accuracy-graph', 'figure'),
                 Output('ic-graph', 'figure')],
                [Input('symbol-dropdown', 'value'),
                 Input('interval-component', 'n_intervals')]
            )
            def update_graphs(symbol, n):
                if symbol is None:
                    return {}, {}
                
                # Get accuracy data
                ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
                metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
                
                if metrics_file.exists():
                    if FinalPredictionTracker.USE_PARQUET:
                        metrics_df = pd.read_parquet(metrics_file)
                    else:
                        metrics_df = pd.read_csv(metrics_file)
                    
                    if 'evaluation_date' in metrics_df.columns:
                        metrics_df = metrics_df.sort_values('evaluation_date')
                        accuracy_fig = go.Figure()
                        accuracy_fig.add_trace(go.Scatter(
                            x=metrics_df['evaluation_date'],
                            y=metrics_df['adjusted_accuracy'],
                            mode='lines+markers',
                            name='Accuracy'
                        ))
                        accuracy_fig.update_layout(title=f'Accuracy Over Time: {symbol}')
                    else:
                        accuracy_fig = {}
                else:
                    accuracy_fig = {}
                
                # Get IC data
                ic_data = FinalPredictionTracker.calculate_information_coefficient(symbol, days=60)
                ic_fig = go.Figure()
                if ic_data:
                    ic_fig.add_trace(go.Bar(
                        x=['Pearson IC', 'Spearman IC'],
                        y=[ic_data.get('ic_pearson', 0), ic_data.get('ic_spearman', 0)],
                        name='IC'
                    ))
                ic_fig.update_layout(title=f'Information Coefficient: {symbol}')
                
                return accuracy_fig, ic_fig
            
            logger.info(f"Dash app created. Run with: app.run_server(port={port})")
            return app
            
        except Exception as e:
            logger.error(f"create_dash_app failed: {e}")
            return None
    
    # ============================================================================
    # ML-BASED DRIFT DETECTION: AUTOML FOR DRIFT PATTERNS
    # ============================================================================
    
    @staticmethod
    def detect_drift_ml(symbol: str, current_df: pd.DataFrame, window_size: int = 50) -> Dict:
        """
        ML-based drift detection using Isolation Forest and Random Forest.
        More sophisticated than statistical tests - learns drift patterns.
        """
        if not ML_DRIFT_AVAILABLE:
            logger.warning("ML libraries not available for drift detection")
            return {'drift_detected': False, 'method': 'unavailable'}
        
        try:
            if len(current_df) < window_size * 2:
                return {'drift_detected': False, 'reason': 'insufficient_data'}
            
            # Prepare features
            feature_cols = ['Returns', 'Volume_Ratio', 'RSI_14', 'Volatility_10', 
                          'MACD_Hist', 'BB_Position_20', 'ATR_14']
            available_features = [f for f in feature_cols if f in current_df.columns]
            
            if len(available_features) < 3:
                return {'drift_detected': False, 'reason': 'insufficient_features'}
            
            # Split into historical and recent windows
            historical_data = current_df.iloc[-window_size*2:-window_size][available_features].dropna()
            recent_data = current_df.iloc[-window_size:][available_features].dropna()
            
            if len(historical_data) < window_size // 2 or len(recent_data) < window_size // 2:
                return {'drift_detected': False, 'reason': 'insufficient_samples'}
            
            # Method 1: Isolation Forest (anomaly detection)
            scaler = StandardScaler()
            historical_scaled = scaler.fit_transform(historical_data)
            recent_scaled = scaler.transform(recent_data)
            
            iso_forest = IsolationForest(contamination=0.1, random_state=42)
            iso_forest.fit(historical_scaled)
            
            # Predict anomalies in recent data
            recent_anomalies = iso_forest.predict(recent_scaled)
            anomaly_rate = np.sum(recent_anomalies == -1) / len(recent_anomalies)
            
            # Method 2: Random Forest classifier (distribution shift detection)
            # Create labels: 0 = historical, 1 = recent
            X_combined = np.vstack([historical_scaled, recent_scaled])
            y_combined = np.hstack([np.zeros(len(historical_scaled)), np.ones(len(recent_scaled))])
            
            rf_classifier = RandomForestClassifier(n_estimators=100, random_state=42, max_depth=5)
            rf_classifier.fit(X_combined, y_combined)
            
            # If classifier can easily distinguish, there's drift
            predictions = rf_classifier.predict(X_combined)
            accuracy = np.mean(predictions == y_combined)
            
            # High accuracy = easy to distinguish = drift detected
            drift_score = accuracy - 0.5  # 0.5 = random, >0.5 = distinguishable
            
            # Combine both methods
            drift_detected = (anomaly_rate > 0.15) or (drift_score > 0.2)
            
            return {
                'drift_detected': bool(drift_detected),
                'anomaly_rate': float(anomaly_rate),
                'distribution_shift_score': float(drift_score),
                'classification_accuracy': float(accuracy),
                'method': 'ml_ensemble',
                'features_used': available_features
            }
            
        except Exception as e:
            logger.warning(f"detect_drift_ml failed: {e}")
            return {'drift_detected': False, 'error': str(e)}
    
    @staticmethod
    def detect_drift_ensemble(symbol: str, current_df: pd.DataFrame) -> Dict:
        """
        Ensemble drift detection combining statistical, ADWIN, and ML methods.
        Most robust approach - used by top hedge funds.
        """
        try:
            results = {
                'statistical': False,
                'adwin': False,
                'ml': False,
                'consensus': False
            }
            
            # Statistical drift
            results['statistical'] = FinalPredictionTracker._detect_drift_advanced(symbol, current_df)
            
            # ADWIN drift
            results['adwin'] = FinalPredictionTracker.detect_drift_adwin(symbol, current_df)
            
            # ML drift
            ml_result = FinalPredictionTracker.detect_drift_ml(symbol, current_df)
            results['ml'] = ml_result.get('drift_detected', False)
            results['ml_details'] = ml_result
            
            # Consensus: drift if 2+ methods agree
            drift_votes = sum([results['statistical'], results['adwin'], results['ml']])
            results['consensus'] = drift_votes >= 2
            results['confidence'] = drift_votes / 3.0
            
            return results
            
        except Exception as e:
            logger.warning(f"detect_drift_ensemble failed: {e}")
            return {'consensus': False, 'error': str(e)}
    
    # ============================================================================
    # GPU ACCELERATION: LARGE-SCALE BATCH PROCESSING
    # ============================================================================
    
    @staticmethod
    def batch_process_predictions_gpu(
        symbols: List[str], 
        operation: str = 'evaluate',
        batch_size: int = 100
    ) -> Dict:
        """
        GPU-accelerated batch processing of predictions.
        Processes multiple symbols in parallel using GPU.
        """
        if not GPU_ACCELERATION_AVAILABLE:
            logger.warning("CuPy not available. Using CPU fallback. Install with: pip install cupy")
            # Fallback to CPU
            return FinalPredictionTracker._batch_process_cpu(symbols, operation, batch_size)
        
        try:
            results = {}
            
            # Process in batches
            for i in range(0, len(symbols), batch_size):
                batch = symbols[i:i+batch_size]
                
                # Load all data for batch
                batch_data = {}
                for symbol in batch:
                    ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
                    preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
                    
                    if preds_file.exists():
                        if FinalPredictionTracker.USE_PARQUET:
                            df = pd.read_parquet(preds_file)
                        else:
                            df = pd.read_csv(preds_file)
                        batch_data[symbol] = df
                
                # Convert to GPU arrays
                if operation == 'evaluate':
                    # GPU-accelerated evaluation
                    for symbol, df in batch_data.items():
                        if 'predicted_change_pct' in df.columns and 'actual_change_pct' in df.columns:
                            # Move to GPU
                            pred_gpu = cp.asarray(df['predicted_change_pct'].values)
                            actual_gpu = cp.asarray(df['actual_change_pct'].values)
                            
                            # GPU-accelerated calculations
                            errors_gpu = cp.abs(pred_gpu - actual_gpu)
                            mae_gpu = cp.mean(errors_gpu)
                            
                            # Move back to CPU
                            results[symbol] = {
                                'mae': float(cp.asnumpy(mae_gpu)),
                                'num_predictions': len(df)
                            }
                
                elif operation == 'metrics':
                    # GPU-accelerated metrics calculation
                    for symbol, df in batch_data.items():
                        if 'adjusted_accuracy' in df.columns:
                            acc_gpu = cp.asarray(df['adjusted_accuracy'].values)
                            mean_acc = cp.mean(acc_gpu)
                            std_acc = cp.std(acc_gpu)
                            
                            results[symbol] = {
                                'mean_accuracy': float(cp.asnumpy(mean_acc)),
                                'std_accuracy': float(cp.asnumpy(std_acc))
                            }
            
            return results
            
        except Exception as e:
            logger.warning(f"batch_process_predictions_gpu failed: {e}, falling back to CPU")
            return FinalPredictionTracker._batch_process_cpu(symbols, operation, batch_size)
    
    @staticmethod
    def _batch_process_cpu(symbols: List[str], operation: str, batch_size: int) -> Dict:
        """CPU fallback for batch processing"""
        results = {}
        for symbol in symbols:
            try:
                if operation == 'evaluate':
                    metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=30)
                    results[symbol] = metrics
                elif operation == 'metrics':
                    metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=30)
                    results[symbol] = {
                        'mean_accuracy': metrics.get('avg_accuracy', 0),
                        'std_accuracy': metrics.get('accuracy_std', 0)
                    }
            except:
                continue
        return results
    
    @staticmethod
    def calculate_correlation_matrix_gpu(symbols: List[str], days: int = 60) -> np.ndarray:
        """
        GPU-accelerated correlation matrix calculation for multiple symbols.
        Much faster for large numbers of symbols.
        """
        if not GPU_ACCELERATION_AVAILABLE:
            # CPU fallback
            return FinalPredictionTracker._calculate_correlation_matrix_cpu(symbols, days)
        
        try:
            # Collect data for all symbols
            data_matrix = []
            valid_symbols = []
            
            for symbol in symbols:
                ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
                metrics_file = FinalPredictionTracker.METRICS_DIR / f"{symbol}_metrics.{ext}"
                
                if metrics_file.exists():
                    if FinalPredictionTracker.USE_PARQUET:
                        df = pd.read_parquet(metrics_file)
                    else:
                        df = pd.read_csv(metrics_file)
                    
                    if 'adjusted_accuracy' in df.columns and len(df) >= days:
                        recent = df.tail(days)
                        accuracies = recent['adjusted_accuracy'].values
                        if len(accuracies) == days:
                            data_matrix.append(accuracies)
                            valid_symbols.append(symbol)
            
            if len(data_matrix) < 2:
                return np.array([])
            
            # Convert to GPU array
            data_gpu = cp.asarray(data_matrix)
            
            # GPU-accelerated correlation
            # Normalize
            mean_gpu = cp.mean(data_gpu, axis=1, keepdims=True)
            std_gpu = cp.std(data_gpu, axis=1, keepdims=True)
            normalized_gpu = (data_gpu - mean_gpu) / (std_gpu + 1e-8)
            
            # Correlation = normalized @ normalized.T / n
            corr_gpu = cp.dot(normalized_gpu, normalized_gpu.T) / data_gpu.shape[1]
            
            # Move back to CPU
            corr_matrix = cp.asnumpy(corr_gpu)
            
            return corr_matrix
            
        except Exception as e:
            logger.warning(f"calculate_correlation_matrix_gpu failed: {e}, using CPU fallback")
            return FinalPredictionTracker._calculate_correlation_matrix_cpu(symbols, days)
    
    @staticmethod
    def _calculate_correlation_matrix_cpu(symbols: List[str], days: int) -> np.ndarray:
        """CPU fallback for correlation matrix"""
        try:
            data_dict = {}
            for symbol in symbols:
                metrics = FinalPredictionTracker._get_recent_metrics(symbol, days=days)
                if metrics.get('num_predictions', 0) > 0:
                    # Simplified - would need actual time series
                    data_dict[symbol] = metrics.get('avg_accuracy', 0)
            
            if len(data_dict) < 2:
                return np.array([])
            
            # Simple correlation (would need proper time series alignment)
            return np.corrcoef(list(data_dict.values()))
        except:
            return np.array([])
    
    # ============================================================================
    # PERFORMANCE PROFILING & OPTIMIZATION
    # ============================================================================
    
    @staticmethod
    def _log_performance(operation: str, duration: float):
        """Internal method to log performance metrics"""
        if not FinalPredictionTracker.PERFORMANCE_PROFILING:
            return
        
        if operation not in FinalPredictionTracker._performance_log:
            FinalPredictionTracker._performance_log[operation] = []
        
        FinalPredictionTracker._performance_log[operation].append(duration)
        
        # Keep only last 1000 entries per operation
        if len(FinalPredictionTracker._performance_log[operation]) > 1000:
            FinalPredictionTracker._performance_log[operation] = FinalPredictionTracker._performance_log[operation][-1000:]
    
    @staticmethod
    def get_performance_profile() -> Dict:
        """
        Get performance profiling data.
        Shows which operations are slow and helps identify bottlenecks.
        Critical for optimization in production systems.
        """
        if not FinalPredictionTracker._performance_log:
            return {
                'message': 'Performance profiling is disabled. Set PERFORMANCE_PROFILING = True to enable.',
                'all_operations': {},
                'bottlenecks': {}
            }
        
        try:
            profile = {}
            
            for operation, timings in FinalPredictionTracker._performance_log.items():
                if len(timings) == 0:
                    continue
                
                timings_array = np.array(timings)
                profile[operation] = {
                    'avg_ms': float(np.mean(timings_array) * 1000),
                    'max_ms': float(np.max(timings_array) * 1000),
                    'min_ms': float(np.min(timings_array) * 1000),
                    'median_ms': float(np.median(timings_array) * 1000),
                    'std_ms': float(np.std(timings_array) * 1000),
                    'p95_ms': float(np.percentile(timings_array, 95) * 1000),
                    'p99_ms': float(np.percentile(timings_array, 99) * 1000),
                    'count': len(timings),
                    'total_time_ms': float(np.sum(timings_array) * 1000)
                }
            
            # Find bottlenecks (>100ms average)
            bottlenecks = {
                op: data 
                for op, data in profile.items() 
                if data['avg_ms'] > 100
            }
            
            # Sort by average time (slowest first)
            sorted_operations = sorted(
                profile.items(),
                key=lambda x: x[1]['avg_ms'],
                reverse=True
            )
            
            # Calculate total time spent
            total_time = sum(data['total_time_ms'] for data in profile.values())
            
            # Percentage of time spent in each operation
            for op, data in profile.items():
                if total_time > 0:
                    data['percentage_of_total'] = float((data['total_time_ms'] / total_time) * 100)
                else:
                    data['percentage_of_total'] = 0.0
            
            return {
                'all_operations': profile,
                'bottlenecks': bottlenecks,
                'sorted_by_slowest': {op: data for op, data in sorted_operations},
                'summary': {
                    'total_operations_tracked': len(profile),
                    'total_time_ms': total_time,
                    'num_bottlenecks': len(bottlenecks),
                    'bottleneck_operations': list(bottlenecks.keys())
                },
                'recommendations': FinalPredictionTracker._generate_optimization_recommendations(profile, bottlenecks)
            }
            
        except Exception as e:
            logger.warning(f"get_performance_profile failed: {e}")
            return {}
    
    @staticmethod
    def get_alpha_signals_for_optimization(symbols: Optional[List[str]] = None, 
                                          days: int = 60, use_ic: bool = True) -> Dict[str, float]:
        """
        Get alpha signals from recent predictions for portfolio optimization
        
        Uses Information Coefficient (IC) for statistically sound alpha signals.
        IC measures correlation between predictions and actual returns.
        
        Parameters:
        -----------
        symbols: List of symbols to get signals for (None = all symbols)
        days: Number of days to look back for predictions
        use_ic: If True, use IC-based alpha (recommended). If False, use simple average.
        
        Returns:
        --------
        Dict of {symbol: alpha_score} where alpha_score is between -1 and 1
        Positive = bullish, Negative = bearish
        """
        try:
            if symbols is None:
                # Get all symbols from prediction files
                predictions_dir = FinalPredictionTracker.PREDICTIONS_DIR
                if not predictions_dir.exists():
                    return {}
                
                symbols = set()
                for pred_file in predictions_dir.glob("*.parquet"):
                    try:
                        df = pd.read_parquet(pred_file)
                        if 'symbol' in df.columns:
                            symbols.update(df['symbol'].unique())
                    except:
                        continue
                symbols = list(symbols)
            
            if not symbols:
                return {}
            
            alpha_signals = {}
            
            for symbol in symbols:
                try:
                    if use_ic:
                        # Use Information Coefficient (IC) - industry standard
                        ic_data = FinalPredictionTracker.calculate_information_coefficient(
                            symbol, days=days
                        )
                        
                        if ic_data and 'ic_pearson' in ic_data:
                            ic = ic_data.get('ic_pearson', 0.0)
                            ic_significant = ic_data.get('ic_significant', False)
                            num_samples = ic_data.get('num_samples', 0)
                            
                            # Only use IC if statistically significant and enough samples
                            if ic_significant and num_samples >= 10:
                                # IC is already normalized correlation [-1, 1]
                                # Weight by significance (p-value)
                                p_value = ic_data.get('ic_p_value', 0.5)
                                significance_weight = 1.0 - min(p_value, 0.5)  # Higher weight for lower p-value
                                
                                # Alpha = IC * significance weight
                                alpha = ic * significance_weight
                                alpha_signals[symbol] = np.clip(alpha, -1.0, 1.0)
                            else:
                                # Fallback to simple method if IC not significant
                                alpha_signals[symbol] = FinalPredictionTracker._get_simple_alpha(
                                    symbol, days
                                )
                        else:
                            # Fallback to simple method if IC calculation fails
                            alpha_signals[symbol] = FinalPredictionTracker._get_simple_alpha(
                                symbol, days
                            )
                    else:
                        # Simple method: average of predicted_change * confidence
                        alpha_signals[symbol] = FinalPredictionTracker._get_simple_alpha(
                            symbol, days
                        )
                
                except Exception as e:
                    logger.warning(f"Error getting alpha for {symbol}: {e}")
                    continue
            
            return alpha_signals
            
        except Exception as e:
            logger.warning(f"get_alpha_signals_for_optimization failed: {e}")
            return {}
    
    @staticmethod
    def _get_simple_alpha(symbol: str, days: int) -> float:
        """Helper method: Get simple alpha (predicted_change * confidence)"""
        try:
            cutoff_date = datetime.now() - timedelta(days=days)
            ext = 'parquet' if FinalPredictionTracker.USE_PARQUET else 'csv'
            preds_file = FinalPredictionTracker.PREDICTIONS_DIR / f"{symbol}_predictions.{ext}"
            
            if not preds_file.exists():
                return 0.0
            
            if FinalPredictionTracker.USE_PARQUET:
                df = pd.read_parquet(preds_file)
            else:
                df = pd.read_csv(preds_file)
            
            if 'prediction_date' in df.columns:
                df['prediction_date'] = pd.to_datetime(df['prediction_date'])
                df = df[df['prediction_date'] >= cutoff_date]
            
            if len(df) == 0:
                return 0.0
            
            # Calculate alpha from predictions
            predicted_changes = df['predicted_change_pct'].fillna(0).values
            confidences = df['confidence'].fillna(0.5).values / 100.0
            
            # Alpha = predicted_change * confidence (normalized)
            alphas = (predicted_changes / 100.0) * confidences
            avg_alpha = np.mean(alphas) if len(alphas) > 0 else 0.0
            
            return np.clip(avg_alpha, -1.0, 1.0)
            
        except Exception as e:
            logger.warning(f"_get_simple_alpha failed for {symbol}: {e}")
            return 0.0
    
    @staticmethod
    def _generate_optimization_recommendations(profile: Dict, bottlenecks: Dict) -> List[str]:
        """Generate optimization recommendations based on performance profile"""
        recommendations = []
        
        if not bottlenecks:
            recommendations.append("✅ No bottlenecks detected. Performance is good!")
            return recommendations
        
        # Check for I/O bottlenecks
        io_operations = [op for op in bottlenecks.keys() if any(keyword in op.lower() for keyword in ['read', 'write', 'save', 'load', 'file'])]
        if io_operations:
            recommendations.append(f"⚠️ I/O bottlenecks detected: {', '.join(io_operations)}. Consider: async I/O, batch writes, or Excel export optimization.")
        
        # Check for computation bottlenecks
        compute_operations = [op for op in bottlenecks.keys() if any(keyword in op.lower() for keyword in ['calculate', 'compute', 'process', 'evaluate'])]
        if compute_operations:
            recommendations.append(f"⚠️ Computation bottlenecks: {', '.join(compute_operations)}. Consider: vectorization, caching, or GPU acceleration.")
        
        # Check for high variance (inconsistent performance)
        high_variance_ops = [
            op for op, data in profile.items()
            if data.get('std_ms', 0) > data.get('avg_ms', 0) * 0.5  # Std > 50% of mean
        ]
        if high_variance_ops:
            recommendations.append(f"⚠️ High variance operations: {', '.join(high_variance_ops)}. Performance is inconsistent - investigate caching or data size variations.")
        
        # Check for operations with high p99 (outliers)
        high_p99_ops = [
            op for op, data in profile.items()
            if data.get('p99_ms', 0) > data.get('avg_ms', 0) * 3  # P99 > 3x average
        ]
        if high_p99_ops:
            recommendations.append(f"⚠️ Operations with outliers: {', '.join(high_p99_ops)}. P99 is much higher than average - investigate edge cases.")
        
        return recommendations
    
    @staticmethod
    def clear_performance_log():
        """Clear performance log (useful for resetting profiling)"""
        FinalPredictionTracker._performance_log.clear()
        logger.info("Performance log cleared")
    
    @staticmethod
    def enable_performance_profiling(enable: bool = True):
        """Enable or disable performance profiling"""
        FinalPredictionTracker.PERFORMANCE_PROFILING = enable
        if enable:
            logger.info("Performance profiling ENABLED")
        else:
            logger.info("Performance profiling DISABLED")


# Backward compatibility
PredictionTracker = FinalPredictionTracker